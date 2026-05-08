"""Парсинг Excel для импорта движений (листы «Поступление» / «Отгрузка» или legacy «movements», ТЗ «Загрузка excel»)."""

from __future__ import annotations

import io
import re
from datetime import date, datetime
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from openpyxl.utils.datetime import from_excel

LEGACY_SHEET_NAME = "movements"


def _sheet_titles_to_try(op_type: str) -> list[str]:
    ot = str(op_type).strip().lower()
    if ot == "in":
        return ["Поступление", LEGACY_SHEET_NAME]
    if ot == "out":
        return ["Отгрузка", LEGACY_SHEET_NAME]
    return ["Поступление", "Отгрузка", LEGACY_SHEET_NAME]


def _pick_movements_worksheet(wb: Workbook, op_type: str) -> Any | None:
    order = _sheet_titles_to_try(op_type)
    lower_map = {str(n).strip().lower(): n for n in wb.sheetnames}
    for cand in order:
        key = str(cand).strip().lower()
        if key in lower_map:
            return wb[lower_map[key]]
    if wb.sheetnames:
        return wb[wb.sheetnames[0]]
    return None

# Канонический ключ -> допустимые заголовки (нижний регистр, без крайних пробелов)
HEADER_SYNONYMS: dict[str, frozenset[str]] = {
    "date": frozenset({"date", "дата"}),
    "name": frozenset({"name", "название"}),
    "barcode": frozenset({"barcode", "штрих-код", "штрихкод", "sku", "шк"}),
    "color": frozenset({"color", "цвет"}),
    "size": frozenset({"size", "размер"}),
    "quantity": frozenset({"quantity", "количество", "qty"}),
    "status": frozenset({"status", "статус"}),
    "comment": frozenset({"comment", "комментарий", "примечание"}),
    "client": frozenset({"client", "клиент"}),
}

# Все обязательные колонки (лист должен содержать заголовки); ТЗ «Импорт excel шаг 02».
REQUIRED_HEADERS = (
    "date",
    "client",
    "name",
    "barcode",
    "color",
    "size",
    "quantity",
    "status",
    "comment",
)

STRUCTURE_ERROR = "Неверная структура файла"


def _norm_header(cell: object) -> str:
    if cell is None:
        return ""
    s = str(cell).strip().lower()
    s = re.sub(r"\s+", " ", s)
    return s


def _map_headers(header_row: list[Any]) -> dict[str, int]:
    """Индекс 0-based колонки по каноническому ключу."""
    idx_by_key: dict[str, int] = {}
    for col_idx, cell in enumerate(header_row):
        h = _norm_header(cell)
        if not h:
            continue
        for key, synonyms in HEADER_SYNONYMS.items():
            if h in synonyms:
                if key not in idx_by_key:
                    idx_by_key[key] = col_idx
                break
    return idx_by_key


def _cell_to_date_str(val: object) -> str | None:
    if val is None or (isinstance(val, str) and not str(val).strip()):
        return None
    if isinstance(val, datetime):
        return val.date().isoformat()
    if isinstance(val, date):
        return val.isoformat()
    if isinstance(val, (int, float)):
        try:
            dt = from_excel(float(val))
            return dt.date().isoformat()
        except Exception:
            return None
    s = str(val).strip()
    if re.fullmatch(r"\d{4}-\d{2}-\d{2}", s):
        return s
    m = re.fullmatch(r"(\d{1,2})\.(\d{1,2})\.(\d{4})", s)
    if m:
        d, mo, y = int(m.group(1)), int(m.group(2)), int(m.group(3))
        try:
            return date(y, mo, d).isoformat()
        except ValueError:
            return None
    return None


def _cell_to_str(val: object) -> str:
    if val is None:
        return ""
    if isinstance(val, float) and val == int(val):
        return str(int(val))
    return str(val).strip()


def normalize_excel_text(val: object) -> str:
    """Трим, схлопывание пробелов (ТЗ шаг 02)."""
    s = _cell_to_str(val)
    if not s:
        return ""
    return re.sub(r"\s+", " ", s).strip()


def _cell_to_int_qty(val: object) -> int | None:
    if val is None:
        return None
    if isinstance(val, bool):
        return None
    if isinstance(val, int):
        return val if val > 0 else None
    if isinstance(val, float):
        if val != int(val):
            return None
        iv = int(val)
        return iv if iv > 0 else None
    s = str(val).strip().replace(",", ".")
    if not s:
        return None
    try:
        n = float(s)
        if n != int(n):
            return None
        iv = int(n)
        return iv if iv > 0 else None
    except ValueError:
        return None


def _is_zip_xlsx(file_bytes: bytes) -> bool:
    return len(file_bytes) >= 4 and file_bytes[0:2] == b"PK"


def _append_movement_row(
    out: list[dict[str, Any]],
    *,
    excel_row: int,
    date_raw: Any,
    name_raw: Any,
    barcode_raw: Any,
    color_raw: Any,
    size_raw: Any,
    qty_raw: Any,
    status_raw: Any,
    comment_raw: Any,
    client_raw: Any,
) -> None:
    date_str = _cell_to_date_str(date_raw)
    name = normalize_excel_text(name_raw)
    barcode = normalize_excel_text(barcode_raw)
    color = normalize_excel_text(color_raw)
    size = normalize_excel_text(size_raw)
    qty = _cell_to_int_qty(qty_raw)
    status_display = normalize_excel_text(status_raw)
    comment = normalize_excel_text(comment_raw)
    client = normalize_excel_text(client_raw)
    q_empty = qty_raw is None or (isinstance(qty_raw, str) and not str(qty_raw).strip())
    out.append(
        {
            "excel_row": excel_row,
            "date": date_str,
            "name": name,
            "barcode": barcode,
            "color": color,
            "size": size or None,
            "quantity_raw": qty_raw,
            "quantity_empty": q_empty,
            "quantity": qty,
            "status_display": status_display,
            "comment": comment or None,
            "client": client or None,
        }
    )


def _parse_movements_openpyxl(file_bytes: bytes, op_type: str) -> tuple[list[dict[str, Any]], list[str]]:
    errors: list[str] = []
    try:
        wb = load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
    except Exception:
        return [], ["Не удалось прочитать файл"]

    ws = _pick_movements_worksheet(wb, op_type)
    if ws is None:
        wb.close()
        return [], ["Не удалось прочитать файл"]

    rows_iter = ws.iter_rows(values_only=True)
    try:
        header_row = next(rows_iter)
    except StopIteration:
        wb.close()
        return [], ["Не удалось прочитать файл"]

    col_map = _map_headers(list(header_row))
    missing = [h for h in REQUIRED_HEADERS if h not in col_map]
    if missing:
        wb.close()
        return [], [STRUCTURE_ERROR]

    out: list[dict[str, Any]] = []
    excel_row = 1
    for data_row in rows_iter:
        excel_row += 1
        if data_row is None or all(v is None or str(v).strip() == "" for v in data_row):
            continue

        def cell(key: str) -> Any:
            j = col_map.get(key)
            if j is None or j >= len(data_row):
                return None
            return data_row[j]

        _append_movement_row(
            out,
            excel_row=excel_row,
            date_raw=cell("date"),
            name_raw=cell("name"),
            barcode_raw=cell("barcode"),
            color_raw=cell("color"),
            size_raw=cell("size"),
            qty_raw=cell("quantity"),
            status_raw=cell("status"),
            comment_raw=cell("comment"),
            client_raw=cell("client"),
        )

    wb.close()
    if not out:
        errors.append("Нет ни одной строки данных под заголовком")
    return out, errors


def _parse_movements_xls(file_bytes: bytes, op_type: str) -> tuple[list[dict[str, Any]], list[str]]:
    errors: list[str] = []
    try:
        import xlrd
    except ImportError:
        return [], ["Не удалось прочитать файл"]

    try:
        book = xlrd.open_workbook(file_contents=file_bytes)
    except Exception:
        return [], ["Не удалось прочитать файл"]

    sheet = None
    lower_names: dict[str, Any] = {}
    for n in book.sheet_names():
        lower_names[str(n).strip().lower()] = book.sheet_by_name(n)
    for cand in _sheet_titles_to_try(op_type):
        key = str(cand).strip().lower()
        if key in lower_names:
            sheet = lower_names[key]
            break
    if sheet is None:
        try:
            sheet = book.sheet_by_index(0)
        except Exception:
            return [], ["Не удалось прочитать файл"]

    if sheet.nrows < 1:
        return [], ["Не удалось прочитать файл"]

    def xls_cell(r: int, c: int) -> Any:
        if c < 0 or c >= sheet.ncols:
            return None
        t = sheet.cell_type(r, c)
        v = sheet.cell_value(r, c)
        if t == xlrd.XL_CELL_DATE:
            try:
                tup = xlrd.xldate_as_tuple(v, book.datemode)
                return date(tup[0], tup[1], tup[2])
            except Exception:
                return v
        return v

    header_row = [xls_cell(0, c) for c in range(sheet.ncols)]
    col_map = _map_headers(header_row)
    missing = [h for h in REQUIRED_HEADERS if h not in col_map]
    if missing:
        return [], [STRUCTURE_ERROR]

    out: list[dict[str, Any]] = []
    for r in range(1, sheet.nrows):
        excel_row = r + 1
        row_vals = [xls_cell(r, c) for c in range(sheet.ncols)]
        if all(v is None or (isinstance(v, str) and not v.strip()) for v in row_vals):
            continue

        def cell(key: str) -> Any:
            j = col_map.get(key)
            if j is None or j >= len(row_vals):
                return None
            return row_vals[j]

        _append_movement_row(
            out,
            excel_row=excel_row,
            date_raw=cell("date"),
            name_raw=cell("name"),
            barcode_raw=cell("barcode"),
            color_raw=cell("color"),
            size_raw=cell("size"),
            qty_raw=cell("quantity"),
            status_raw=cell("status"),
            comment_raw=cell("comment"),
            client_raw=cell("client"),
        )

    if not out:
        errors.append("Нет ни одной строки данных под заголовком")
    return out, errors


def parse_movements_excel(file_bytes: bytes, op_type: str) -> tuple[list[dict[str, Any]], list[str]]:
    """Читает файл; возвращает (строки данных, ошибки файла/структуры).

    Строка: ``excel_row``, ``date`` (ISO или None), текстовые поля нормализованы,
    ``status_display`` — как в файле, ``quantity`` — int или None.
    """
    if not file_bytes or len(file_bytes) < 10:
        return [], ["Не удалось прочитать файл"]

    if _is_zip_xlsx(file_bytes):
        return _parse_movements_openpyxl(file_bytes, op_type)
    return _parse_movements_xls(file_bytes, op_type)


def normalize_movement_import_status(status_raw: str, op_type: str) -> str | None:
    """Текст колонки «Статус» → ``planned`` | ``accepted`` для логики импорта."""
    s = normalize_excel_text(status_raw)
    if not s:
        return None
    sf = s.casefold()
    if sf in ("planned", "pending", "draft"):
        return "planned"
    if sf in ("accepted", "shipped"):
        return "accepted"
    ot = str(op_type).strip().lower()
    if ot == "in":
        if sf in (
            "ожидание приемки",
            "ожидает приемки",
            "ожидает приёмки",
            "ожидает поступления",
            "ожидает поступления для приемки",
            "ожидает поступления для приёмки",
        ):
            return "planned"
        if sf in ("принят", "принято", "принята", "приняты"):
            return "accepted"
    elif ot == "out":
        if sf in ("ожидание отгрузки", "ожидает отгрузки"):
            return "planned"
        if sf in ("отгружен", "отгружено", "отгружена", "отгружены"):
            return "accepted"
    return None


def _excel_column_width_from_px(px: int) -> float:
    """Приближение ширины столбца в символах Excel к целевой ширине в пикселях (Calibri по умолчанию)."""
    return max(2.0, round((px - 12) / 7 + 1, 2))


def build_template_workbook_bytes(op_type: str) -> bytes:
    """Генерирует xlsx-шаблон с листом «Поступление» / «Отгрузка», русскими заголовками и примером строки (дата дд.мм.гггг)."""
    wb = Workbook()
    ws = wb.active
    assert ws is not None
    ot = str(op_type).strip().lower()
    ws.title = "Поступление" if ot == "in" else "Отгрузка"
    headers = ["Дата", "Клиент", "Название", "ШК", "Цвет", "Размер", "Количество", "Статус", "Комментарий"]
    ws.append(headers)
    bold_header = Font(bold=True)
    for cell in ws[1]:
        cell.font = bold_header

    status_example = "Ожидает поступления для приемки" if ot == "in" else "Ожидает отгрузки"
    ws.append(
        [
            "01.01.1991",
            "Рога и копыта",
            "Водолазка",
            "123456789",
            "Черный",
            "44",
            500,
            status_example,
            '="ДАННЫЕ ДЛЯ ПРИМЕРА, УДАЛИТЕ СТРОКУ"',
        ]
    )
    pixel_widths = [75, 120, 120, 120, 120, 70, 85, 150, 300]
    for idx, px in enumerate(pixel_widths, start=1):
        col = get_column_letter(idx)
        ws.column_dimensions[col].width = _excel_column_width_from_px(px)
    bio = io.BytesIO()
    wb.save(bio)
    return bio.getvalue()
