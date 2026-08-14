from __future__ import annotations

from io import BytesIO

from fastapi import HTTPException, status
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

from config import (
    CABINET_DISPATCH_OPS_VISIBLE,
    CABINET_DISPATCH_VISIBLE_STATUSES,
    CABINET_RECEIPT_OPS_VISIBLE,
    CABINET_RECEIPT_VISIBLE_STATUSES,
    DISPATCH_CARGO_GOOD,
    DISPATCH_CARGO_TYPES,
    DISPATCH_STATUS_AWAITING_TRIP,
    DISPATCH_STATUS_PARTIALLY_SHIPPED,
    DISPATCH_STATUS_PREPARING,
    INV_OP_INTAKE,
    INV_OP_SINKS,
    INV_OP_WRITTEN_OFF,
    INV_Q_DEFECT,
    INV_Q_GOOD,
    PRODUCT_LIST_SORT_COLUMNS,
    RECEIPT_STATUS_PARTIALLY_RECEIVED,
    RECEIPT_STATUS_PLANNED,
    TRIP_STATUS_CANCELLED,
)
from dbconn import ci_like_substring_param
from modules.balances.service import get_balances
from modules.timesheet.service import business_today
from modules.cabinet.schemas import (
    CabinetBalanceTotals,
    CabinetEventItem,
    CabinetOpItem,
    CabinetPackingReportDay,
    CabinetPackingReportResponse,
    CabinetPackingReportRow,
    CabinetProfileResponse,
    CabinetReceiptDetailResponse,
    CabinetReceiptDoc,
    CabinetReceiptLine,
    CabinetReceiptLineItem,
    CabinetReceiptLinesResponse,
    CabinetReceiptListItem,
    CabinetReceiptListResponse,
    CabinetReceiptTotals,
    CabinetShipmentDetailResponse,
    CabinetShipmentDoc,
    CabinetShipmentLine,
    CabinetShipmentLineItem,
    CabinetShipmentLinesResponse,
    CabinetShipmentListItem,
    CabinetShipmentListResponse,
    CabinetShipmentTrip,
    CabinetSummaryResponse,
    CabinetWriteOffItem,
    CabinetWriteOffsResponse,
)
from modules.products.schemas import ProductItem, ProductListResponse, ProductVariantDimension, ProductVariantItem
from modules.products.service import (
    _ci_substring_like_param,
    _decode_images_json,
    _order_sql_from_sort_param,
    _row_to_product_item,
)

_SINKS_SQL = ", ".join(f"'{s}'" for s in INV_OP_SINKS)


def list_cabinet_products(
    connection,
    *,
    client_id: str,
    page: int,
    limit: int,
    search: str | None,
    sort: str | None,
) -> ProductListResponse:
    offset = (page - 1) * limit
    conds = ["p.client_id = ?", "COALESCE(p.is_deleted, 0) = 0"]
    params: list[object] = [client_id]
    if search is not None and str(search).strip():
        like = _ci_substring_like_param(str(search))
        conds.append("(fold_ci(COALESCE(p.name, '')) LIKE ? OR fold_ci(COALESCE(p.sku, '')) LIKE ?)")
        params.extend([like, like])

    join_sql = """
        FROM products p
        LEFT JOIN product_types pt ON pt.id = p.type_id
        LEFT JOIN clients c ON c.id = p.client_id
        LEFT JOIN (
            SELECT product_id, COUNT(*) AS cnt
            FROM product_variants
            WHERE COALESCE(is_deleted, 0) = 0
            GROUP BY product_id
        ) vcnt ON vcnt.product_id = p.id
        LEFT JOIN users creator ON creator.id = p.creator_id
        LEFT JOIN users editor ON editor.id = p.updated_by_id
        LEFT JOIN users deleter ON deleter.id = p.deleted_by_id
    """
    where_sql = " AND ".join(conds)
    order_sql = _order_sql_from_sort_param(sort, PRODUCT_LIST_SORT_COLUMNS) or "p.name ASC, p.sku ASC"

    total = int(
        connection.execute(
            f"SELECT COUNT(*) AS cnt {join_sql} WHERE {where_sql}",
            params,
        ).fetchone()["cnt"]
    )
    rows = connection.execute(
        f"""
        SELECT p.id, p.name, p.type_id, pt.name AS type_name, p.sku AS sku_base, p.weight_grams,
               p.items_per_box, p.boxes_per_pallet,
               COALESCE(pt.requires_color, 0) AS requires_color,
               COALESCE(pt.requires_size, 0) AS requires_size,
               p.client_id, c.name AS client_name,
               COALESCE(vcnt.cnt, 0) AS variant_count,
               p.is_active, COALESCE(p.is_deleted, 0) AS is_deleted,
               p.deleted_at, p.image_url, p.gallery_json,
               p.created_at, p.updated_at,
               COALESCE(NULLIF(creator.display_name, ''), creator.email) AS created_by, COALESCE(NULLIF(editor.display_name, ''), editor.email) AS updated_by, COALESCE(NULLIF(deleter.display_name, ''), deleter.email) AS deleted_by
        {join_sql}
        WHERE {where_sql}
        ORDER BY {order_sql}
        LIMIT ? OFFSET ?
        """,
        [*params, limit, offset],
    ).fetchall()
    return ProductListResponse(
        items=[_row_to_product_item(row) for row in rows],
        total=total,
        page=page,
        limit=limit,
    )


def get_cabinet_product(connection, *, client_id: str, product_id: str) -> ProductItem:
    row = connection.execute(
        """
        SELECT p.id, p.name, p.type_id, pt.name AS type_name, p.sku AS sku_base, p.weight_grams,
               p.items_per_box, p.boxes_per_pallet,
               COALESCE(pt.requires_color, 0) AS requires_color,
               COALESCE(pt.requires_size, 0) AS requires_size,
               p.client_id, c.name AS client_name,
               COALESCE(vcnt.cnt, 0) AS variant_count,
               p.is_active, COALESCE(p.is_deleted, 0) AS is_deleted,
               p.deleted_at, p.image_url, p.gallery_json,
               p.created_at, p.updated_at,
               COALESCE(NULLIF(creator.display_name, ''), creator.email) AS created_by, COALESCE(NULLIF(editor.display_name, ''), editor.email) AS updated_by, COALESCE(NULLIF(deleter.display_name, ''), deleter.email) AS deleted_by
        FROM products p
        LEFT JOIN product_types pt ON pt.id = p.type_id
        LEFT JOIN clients c ON c.id = p.client_id
        LEFT JOIN (
            SELECT product_id, COUNT(*) AS cnt
            FROM product_variants
            WHERE COALESCE(is_deleted, 0) = 0
            GROUP BY product_id
        ) vcnt ON vcnt.product_id = p.id
        LEFT JOIN users creator ON creator.id = p.creator_id
        LEFT JOIN users editor ON editor.id = p.updated_by_id
        LEFT JOIN users deleter ON deleter.id = p.deleted_by_id
        WHERE p.id = ?
          AND p.client_id = ?
          AND COALESCE(p.is_deleted, 0) = 0
        """,
        (product_id, client_id),
    ).fetchone()
    if not row:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Товар не найден")
    return _row_to_product_item(row)


def list_cabinet_product_variants(connection, *, client_id: str, product_id: str) -> list[ProductVariantItem]:
    product = connection.execute(
        """
        SELECT id
        FROM products
        WHERE id = ?
          AND client_id = ?
          AND COALESCE(is_deleted, 0) = 0
        """,
        (product_id, client_id),
    ).fetchone()
    if not product:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Товар не найден")

    rows = connection.execute(
        f"""
        SELECT v.id, v.color_id, col.name AS color_name,
               v.size_id, sz.name AS size_name,
               v.length, v.width, v.height, v.sku, v.images_json, v.is_active,
               GREATEST(0, COALESCE(b.good_in, 0)) AS stock,
               GREATEST(0, COALESCE(b.defect_in, 0)) AS defect_qty,
               CASE WHEN EXISTS (
                   SELECT 1 FROM receipt_lines rl
                   JOIN receipt_docs rd ON rd.id = rl.doc_id
                   WHERE rl.product_id = v.product_id
                     AND rl.color_id IS NOT DISTINCT FROM v.color_id
                     AND rl.size_id IS NOT DISTINCT FROM v.size_id
                     AND rd.client_id = ?
                     AND rl.is_deleted = 0 AND rd.is_deleted = 0
               ) THEN 1 ELSE 0 END AS has_receipts
        FROM product_variants v
        LEFT JOIN colors col ON col.id = v.color_id
        LEFT JOIN sizes sz ON sz.id = v.size_id
        LEFT JOIN (
            SELECT product_id, color_id, size_id,
                   SUM(CASE WHEN to_quality='{INV_Q_GOOD}' AND to_op NOT IN ({_SINKS_SQL}) THEN qty ELSE 0 END)
                     - SUM(CASE WHEN from_quality='{INV_Q_GOOD}' AND from_op<>'{INV_OP_INTAKE}' THEN qty ELSE 0 END) AS good_in,
                   SUM(CASE WHEN to_quality='{INV_Q_DEFECT}' AND to_op NOT IN ({_SINKS_SQL}) THEN qty ELSE 0 END)
                     - SUM(CASE WHEN from_quality='{INV_Q_DEFECT}' AND from_op<>'{INV_OP_INTAKE}' THEN qty ELSE 0 END) AS defect_in
            FROM zone_relocations
            WHERE product_id = ? AND client_id = ?
            GROUP BY product_id, color_id, size_id
        ) b ON b.product_id = v.product_id AND b.color_id IS NOT DISTINCT FROM v.color_id AND b.size_id IS NOT DISTINCT FROM v.size_id
        WHERE v.product_id = ?
          AND v.client_id = ?
          AND COALESCE(v.is_deleted, 0) = 0
        ORDER BY LOWER(v.sku) ASC
        """,
        (
            client_id,
            product_id,
            client_id,
            product_id,
            client_id,
        ),
    ).fetchall()
    return [
        ProductVariantItem(
            id=str(row["id"]),
            color_id=row["color_id"],
            color_name=row["color_name"],
            dimension=ProductVariantDimension(
                length=float(row["length"]),
                width=float(row["width"]),
                height=float(row["height"]),
            ),
            size_id=str(row["size_id"]) if row["size_id"] else None,
            size_name=row["size_name"],
            sku=str(row["sku"]),
            images=_decode_images_json(row["images_json"]),
            is_active=bool(row["is_active"]),
            stock=max(0, int(row["stock"])),
            defect_qty=max(0, int(row["defect_qty"])),
            has_receipts=bool(row["has_receipts"]),
        )
        for row in rows
    ]


# ---------------------------------------------------------------------------
# Поступления (клиентская проекция)
# ---------------------------------------------------------------------------

def _receipt_status_cond(status: str | None) -> tuple[str, list[str]]:
    allowed = [status] if status else sorted(CABINET_RECEIPT_VISIBLE_STATUSES)
    placeholders = ",".join("?" for _ in allowed)
    return f"d.status IN ({placeholders})", allowed


def _shipment_status_cond(status: str | None) -> tuple[str, list[str]]:
    allowed = [status] if status else sorted(CABINET_DISPATCH_VISIBLE_STATUSES)
    placeholders = ",".join("?" for _ in allowed)
    return f"d.status IN ({placeholders})", allowed


def list_cabinet_receipts(
    connection,
    *,
    client_id: str,
    page: int,
    limit: int,
    status: str | None = None,
    statuses: list[str] | None = None,
    search: str | None = None,
    date_from: str | None = None,
    date_to: str | None = None,
    order: str = "recent",
) -> CabinetReceiptListResponse:
    conds = ["d.is_deleted = 0", "d.client_id = ?"]
    params: list = [client_id]
    if statuses:
        placeholders = ",".join("?" for _ in statuses)
        conds.append(f"d.status IN ({placeholders})")
        params.extend(statuses)
    else:
        cond, status_params = _receipt_status_cond(status)
        conds.append(cond)
        params.extend(status_params)
    if search and search.strip():
        s = ci_like_substring_param(search)
        conds.append("fold_ci(d.doc_number) LIKE ?")
        params.append(s)
    if date_from:
        conds.append("d.arrival_date >= ?")
        params.append(date_from)
    if date_to:
        conds.append("d.arrival_date <= ?")
        params.append(date_to)
    where = " AND ".join(conds)

    total = int(connection.execute(
        f"SELECT COUNT(*) AS cnt FROM receipt_docs d WHERE {where}", params
    ).fetchone()["cnt"])

    order_sql = (
        "d.arrival_date ASC NULLS LAST, d.created_at"
        if order == "upcoming"
        else "COALESCE(d.actual_arrival_date, d.arrival_date) DESC, d.created_at DESC"
    )
    offset = (page - 1) * limit
    rows = connection.execute(
        f"""
        SELECT
            d.id, d.doc_number, d.arrival_date, d.actual_arrival_date,
            d.status, d.created_at,
            COUNT(DISTINCT CASE WHEN l.is_deleted = 0 THEN l.id END) AS sku_count,
            COALESCE(SUM(CASE WHEN l.is_deleted = 0 THEN l.planned_qty ELSE 0 END), 0) AS total_planned,
            COALESCE(SUM(CASE WHEN l.is_deleted = 0 THEN COALESCE(l.accepted_qty, 0) ELSE 0 END), 0) AS total_accepted_qty
        FROM receipt_docs d
        LEFT JOIN receipt_lines l ON l.doc_id = d.id
        WHERE {where}
        GROUP BY d.id
        ORDER BY {order_sql}
        LIMIT ? OFFSET ?
        """,
        params + [limit, offset],
    ).fetchall()

    items = [
        CabinetReceiptListItem(
            id=str(r["id"]),
            doc_number=str(r["doc_number"]),
            arrival_date=r["arrival_date"],
            actual_arrival_date=r["actual_arrival_date"],
            status=str(r["status"]),
            sku_count=int(r["sku_count"] or 0),
            total_planned=int(r["total_planned"] or 0),
            total_accepted_qty=int(r["total_accepted_qty"] or 0),
            created_at=str(r["created_at"]),
        )
        for r in rows
    ]
    return CabinetReceiptListResponse(items=items, total=total, page=page, limit=limit)


def list_cabinet_receipt_lines(
    connection,
    *,
    client_id: str,
    page: int,
    limit: int,
    status: str | None = None,
    search: str | None = None,
    date_from: str | None = None,
    date_to: str | None = None,
) -> CabinetReceiptLinesResponse:
    conds = ["d.is_deleted = 0", "l.is_deleted = 0", "d.client_id = ?"]
    params: list = [client_id]
    cond, status_params = _receipt_status_cond(status)
    conds.append(cond)
    params.extend(status_params)
    if search and search.strip():
        s = ci_like_substring_param(search)
        conds.append("(fold_ci(l.product_sku) LIKE ? OR fold_ci(l.product_name) LIKE ? OR fold_ci(d.doc_number) LIKE ?)")
        params += [s, s, s]
    if date_from:
        conds.append("d.arrival_date >= ?")
        params.append(date_from)
    if date_to:
        conds.append("d.arrival_date <= ?")
        params.append(date_to)
    where = " AND ".join(conds)

    total = int(connection.execute(
        f"""SELECT COUNT(*) AS cnt
            FROM receipt_lines l JOIN receipt_docs d ON d.id = l.doc_id
            WHERE {where}""",
        params,
    ).fetchone()["cnt"])

    offset = (page - 1) * limit
    rows = connection.execute(
        f"""
        SELECT l.doc_id, d.doc_number, d.status, d.arrival_date, d.actual_arrival_date,
               l.product_name, l.product_sku, l.color_name, l.size_name,
               l.planned_qty, l.accepted_qty
        FROM receipt_lines l
        JOIN receipt_docs d ON d.id = l.doc_id
        WHERE {where}
        ORDER BY COALESCE(d.actual_arrival_date, d.arrival_date) DESC, d.created_at DESC, l.created_at
        LIMIT ? OFFSET ?
        """,
        params + [limit, offset],
    ).fetchall()

    items = [
        CabinetReceiptLineItem(
            doc_id=str(r["doc_id"]),
            doc_number=str(r["doc_number"]),
            status=str(r["status"]),
            arrival_date=r["arrival_date"],
            actual_arrival_date=r["actual_arrival_date"],
            product_name=str(r["product_name"]),
            product_sku=str(r["product_sku"]),
            color_name=r["color_name"],
            size_name=r["size_name"],
            planned_qty=int(r["planned_qty"]),
            accepted_qty=int(r["accepted_qty"]) if r["accepted_qty"] is not None else None,
        )
        for r in rows
    ]
    return CabinetReceiptLinesResponse(items=items, total=total, page=page, limit=limit)


def get_cabinet_receipt(connection, *, client_id: str, doc_id: str) -> CabinetReceiptDetailResponse:
    cond, status_params = _receipt_status_cond(None)
    row = connection.execute(
        f"SELECT * FROM receipt_docs d WHERE d.id = ? AND d.client_id = ? AND d.is_deleted = 0 AND {cond}",
        [doc_id, client_id, *status_params],
    ).fetchone()
    if not row:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Документ не найден")

    lines_rows = connection.execute(
        """SELECT product_name, product_sku, color_name, size_name, planned_qty, accepted_qty
           FROM receipt_lines WHERE doc_id = ? AND is_deleted = 0 ORDER BY created_at, id""",
        (doc_id,),
    ).fetchall()
    ops_placeholders = ",".join("?" for _ in CABINET_RECEIPT_OPS_VISIBLE)
    ops_rows = connection.execute(
        f"""SELECT op_type, qty, comment, created_at
            FROM receipt_ops WHERE doc_id = ? AND op_type IN ({ops_placeholders})
            ORDER BY created_at DESC""",
        [doc_id, *sorted(CABINET_RECEIPT_OPS_VISIBLE)],
    ).fetchall()

    lines = [
        CabinetReceiptLine(
            product_name=str(l["product_name"]),
            product_sku=str(l["product_sku"]),
            color_name=l["color_name"],
            size_name=l["size_name"],
            planned_qty=int(l["planned_qty"]),
            accepted_qty=int(l["accepted_qty"]) if l["accepted_qty"] is not None else None,
        )
        for l in lines_rows
    ]
    return CabinetReceiptDetailResponse(
        doc=CabinetReceiptDoc(
            id=str(row["id"]),
            doc_number=str(row["doc_number"]),
            arrival_date=row["arrival_date"],
            actual_arrival_date=row["actual_arrival_date"],
            ttn=row["ttn"],
            status=str(row["status"]),
            created_at=str(row["created_at"]),
        ),
        lines=lines,
        ops=[
            CabinetOpItem(
                op_type=str(o["op_type"]),
                qty=int(o["qty"]) if o["qty"] is not None else None,
                comment=o["comment"],
                created_at=str(o["created_at"]),
            )
            for o in ops_rows
        ],
        totals=CabinetReceiptTotals(
            total_planned=sum(l.planned_qty for l in lines),
            total_accepted=sum(l.accepted_qty or 0 for l in lines),
        ),
    )


# ---------------------------------------------------------------------------
# Отгрузки клиента (домен dispatch, клиентская проекция)
# ---------------------------------------------------------------------------

def list_cabinet_shipments(
    connection,
    *,
    client_id: str,
    page: int,
    limit: int,
    status: str | None = None,
    statuses: list[str] | None = None,
    cargo_type: str | None = None,
    search: str | None = None,
    date_from: str | None = None,
    date_to: str | None = None,
    order: str = "recent",
) -> CabinetShipmentListResponse:
    conds = ["d.is_deleted = 0", "d.client_id = ?"]
    params: list = [client_id]
    if statuses:
        placeholders = ",".join("?" for _ in statuses)
        conds.append(f"d.status IN ({placeholders})")
        params.extend(statuses)
    else:
        cond, status_params = _shipment_status_cond(status)
        conds.append(cond)
        params.extend(status_params)
    if cargo_type in DISPATCH_CARGO_TYPES:
        conds.append(f"COALESCE(d.cargo_type, '{DISPATCH_CARGO_GOOD}') = ?")
        params.append(cargo_type)
    if search and search.strip():
        s = ci_like_substring_param(search)
        conds.append(
            "(fold_ci(d.doc_number) LIKE ? OR EXISTS (SELECT 1 FROM dispatch_lines l"
            " WHERE l.doc_id = d.id AND COALESCE(l.is_deleted, 0) = 0 AND fold_ci(l.store_name) LIKE ?))"
        )
        params += [s, s]
    if date_from:
        conds.append("d.ship_date >= ?")
        params.append(date_from)
    if date_to:
        conds.append("d.ship_date <= ?")
        params.append(date_to)
    where = " AND ".join(conds)

    total = int(connection.execute(
        f"SELECT COUNT(*) AS cnt FROM dispatch_docs d WHERE {where}", params
    ).fetchone()["cnt"])

    order_sql = (
        "d.ship_date ASC NULLS LAST, d.created_at"
        if order == "upcoming"
        else "d.ship_date DESC NULLS LAST, d.created_at DESC"
    )
    offset = (page - 1) * limit
    rows = connection.execute(
        f"""
        SELECT d.id, d.doc_number, COALESCE(d.cargo_type, '{DISPATCH_CARGO_GOOD}') AS cargo_type,
               d.ship_date, d.actual_ship_date,
               d.status, d.created_at,
               ARRAY_REMOVE(ARRAY_AGG(DISTINCT l.store_name) FILTER (WHERE l.is_deleted = 0), NULL) AS store_names,
               COUNT(l.id) FILTER (WHERE l.is_deleted = 0) AS sku_count,
               COALESCE(SUM(l.qty) FILTER (WHERE l.is_deleted = 0), 0) AS total_qty,
               COALESCE(SUM(COALESCE(l.shipped_qty, 0)) FILTER (WHERE l.is_deleted = 0), 0) AS total_shipped_qty
        FROM dispatch_docs d
        LEFT JOIN dispatch_lines l ON l.doc_id = d.id
        WHERE {where}
        GROUP BY d.id
        ORDER BY {order_sql}
        LIMIT ? OFFSET ?
        """,
        params + [limit, offset],
    ).fetchall()

    items = [
        CabinetShipmentListItem(
            id=str(r["id"]),
            doc_number=str(r["doc_number"]),
            cargo_type=str(r["cargo_type"]),
            store_names=[str(s) for s in (r["store_names"] or [])],
            ship_date=r["ship_date"],
            actual_ship_date=r["actual_ship_date"],
            status=str(r["status"]),
            sku_count=int(r["sku_count"] or 0),
            total_qty=int(r["total_qty"] or 0),
            total_shipped_qty=int(r["total_shipped_qty"] or 0),
            created_at=str(r["created_at"]),
        )
        for r in rows
    ]
    return CabinetShipmentListResponse(items=items, total=total, page=page, limit=limit)


def list_cabinet_shipment_lines(
    connection,
    *,
    client_id: str,
    page: int,
    limit: int,
    status: str | None = None,
    cargo_type: str | None = None,
    search: str | None = None,
    date_from: str | None = None,
    date_to: str | None = None,
) -> CabinetShipmentLinesResponse:
    conds = ["d.is_deleted = 0", "l.is_deleted = 0", "d.client_id = ?"]
    params: list = [client_id]
    cond, status_params = _shipment_status_cond(status)
    conds.append(cond)
    params.extend(status_params)
    if cargo_type in DISPATCH_CARGO_TYPES:
        conds.append(f"COALESCE(d.cargo_type, '{DISPATCH_CARGO_GOOD}') = ?")
        params.append(cargo_type)
    if search and search.strip():
        s = ci_like_substring_param(search)
        conds.append("(fold_ci(l.product_sku) LIKE ? OR fold_ci(l.product_name) LIKE ? OR fold_ci(d.doc_number) LIKE ?)")
        params += [s, s, s]
    if date_from:
        conds.append("d.ship_date >= ?")
        params.append(date_from)
    if date_to:
        conds.append("d.ship_date <= ?")
        params.append(date_to)
    where = " AND ".join(conds)

    total = int(connection.execute(
        f"""SELECT COUNT(*) AS cnt
            FROM dispatch_lines l JOIN dispatch_docs d ON d.id = l.doc_id
            WHERE {where}""",
        params,
    ).fetchone()["cnt"])

    offset = (page - 1) * limit
    rows = connection.execute(
        f"""
        SELECT l.doc_id, d.doc_number, COALESCE(d.cargo_type, '{DISPATCH_CARGO_GOOD}') AS cargo_type,
               d.status, d.ship_date,
               l.product_name, l.product_sku, l.color_name, l.size_name,
               l.qty, l.shipped_qty, l.site_url, l.store_name
        FROM dispatch_lines l
        JOIN dispatch_docs d ON d.id = l.doc_id
        WHERE {where}
        ORDER BY d.ship_date DESC NULLS LAST, d.created_at DESC, l.created_at
        LIMIT ? OFFSET ?
        """,
        params + [limit, offset],
    ).fetchall()

    items = [
        CabinetShipmentLineItem(
            doc_id=str(r["doc_id"]),
            doc_number=str(r["doc_number"]),
            cargo_type=str(r["cargo_type"]),
            status=str(r["status"]),
            ship_date=r["ship_date"],
            product_name=str(r["product_name"]),
            product_sku=str(r["product_sku"]),
            color_name=r["color_name"],
            size_name=r["size_name"],
            qty=int(r["qty"] or 0),
            shipped_qty=int(r["shipped_qty"] or 0),
            site_url=r["site_url"],
            store_name=r["store_name"],
        )
        for r in rows
    ]
    return CabinetShipmentLinesResponse(items=items, total=total, page=page, limit=limit)


def list_cabinet_write_offs(
    connection,
    *,
    client_id: str,
    page: int,
    limit: int,
) -> CabinetWriteOffsResponse:
    """Списания товара клиента: журнальные движения → written_off."""
    total = int(connection.execute(
        "SELECT COUNT(*) AS cnt FROM zone_relocations WHERE client_id = ? AND to_op = ?",
        (client_id, INV_OP_WRITTEN_OFF),
    ).fetchone()["cnt"])

    offset = (page - 1) * limit
    rows = connection.execute(
        """
        SELECT id, created_at, product_name, product_sku, color_name, size_name,
               from_quality, qty, reason, comment
        FROM zone_relocations
        WHERE client_id = ? AND to_op = ?
        ORDER BY created_at DESC
        LIMIT ? OFFSET ?
        """,
        (client_id, INV_OP_WRITTEN_OFF, limit, offset),
    ).fetchall()

    items = [
        CabinetWriteOffItem(
            id=str(r["id"]),
            created_at=str(r["created_at"]),
            product_name=r["product_name"],
            product_sku=r["product_sku"],
            color_name=r["color_name"],
            size_name=r["size_name"],
            quality=str(r["from_quality"]),
            qty=int(r["qty"] or 0),
            reason=r["reason"],
            comment=r["comment"],
        )
        for r in rows
    ]
    return CabinetWriteOffsResponse(items=items, total=total, page=page, limit=limit)


def get_cabinet_shipment(connection, *, client_id: str, doc_id: str) -> CabinetShipmentDetailResponse:
    cond, status_params = _shipment_status_cond(None)
    row = connection.execute(
        f"SELECT * FROM dispatch_docs d WHERE d.id = ? AND d.client_id = ? AND d.is_deleted = 0 AND {cond}",
        [doc_id, client_id, *status_params],
    ).fetchone()
    if not row:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Документ не найден")

    lines_rows = connection.execute(
        "SELECT * FROM dispatch_lines WHERE doc_id = ? AND is_deleted = 0 ORDER BY created_at, id",
        (doc_id,),
    ).fetchall()

    trip_rows = connection.execute(
        "SELECT DISTINCT t.id AS trip_id, t.trip_number AS trip_number "
        "FROM trip_lines tl "
        "JOIN trip_docs t ON t.id = tl.trip_id AND COALESCE(t.is_deleted, 0) = 0 "
        "WHERE tl.dispatch_doc_id = ? AND COALESCE(tl.is_deleted, 0) = 0 AND t.status != ? "
        "ORDER BY t.trip_number",
        (doc_id, TRIP_STATUS_CANCELLED),
    ).fetchall()

    ops_placeholders = ",".join("?" for _ in CABINET_DISPATCH_OPS_VISIBLE)
    ops_rows = connection.execute(
        f"""SELECT op_type, comment, created_at
            FROM dispatch_ops WHERE doc_id = ? AND op_type IN ({ops_placeholders})
            ORDER BY created_at DESC""",
        [doc_id, *sorted(CABINET_DISPATCH_OPS_VISIBLE)],
    ).fetchall()

    return CabinetShipmentDetailResponse(
        doc=CabinetShipmentDoc(
            id=str(row["id"]),
            doc_number=str(row["doc_number"]),
            cargo_type=str(row["cargo_type"] or "good"),
            ship_date=row["ship_date"],
            actual_ship_date=row["actual_ship_date"],
            status=str(row["status"]),
            created_at=str(row["created_at"]),
        ),
        lines=[
            CabinetShipmentLine(
                id=str(l["id"]),
                product_name=str(l["product_name"]),
                product_sku=str(l["product_sku"]),
                color_name=l["color_name"],
                size_name=l["size_name"],
                qty=int(l["qty"] or 0),
                shipped_qty=int(l["shipped_qty"] or 0),
                site_url=l["site_url"],
                store_name=l["store_name"],
            )
            for l in lines_rows
        ],
        ops=[
            CabinetOpItem(
                op_type=str(o["op_type"]),
                comment=o["comment"],
                created_at=str(o["created_at"]),
            )
            for o in ops_rows
        ],
        trips=[CabinetShipmentTrip(id=str(t["trip_id"]), number=str(t["trip_number"])) for t in trip_rows],
    )


# ---------------------------------------------------------------------------
# Сводка, отчёты, профиль
# ---------------------------------------------------------------------------

def cabinet_balance_totals(connection, *, client_id: str) -> CabinetBalanceTotals:
    from modules.balances.service import get_balances_summary

    summary = get_balances_summary(connection, client_id=client_id, search=None, has_defect=False)
    # «Упаковано» (packed) для клиента — часть процесса упаковки: сворачиваем в
    # packing_good, чтобы разбивка сходилась с total_good и не плодить новую корзину.
    return CabinetBalanceTotals(
        storage_good=summary.storage_good,
        packing_good=summary.packing_good + summary.packed_good,
        ready_good=summary.ready_good,
        total_good=summary.storage_good + summary.packing_good + summary.packed_good + summary.ready_good,
        defect_total=summary.storage_defect + summary.packing_defect + summary.packed_defect + summary.ready_defect,
    )


def cabinet_summary(connection, *, client_id: str) -> CabinetSummaryResponse:
    active_receipts = list_cabinet_receipts(
        connection,
        client_id=client_id,
        page=1,
        limit=5,
        statuses=[RECEIPT_STATUS_PLANNED, RECEIPT_STATUS_PARTIALLY_RECEIVED],
        order="upcoming",
    ).items
    active_shipments = list_cabinet_shipments(
        connection,
        client_id=client_id,
        page=1,
        limit=5,
        statuses=[
            DISPATCH_STATUS_PREPARING,
            DISPATCH_STATUS_AWAITING_TRIP,
            DISPATCH_STATUS_PARTIALLY_SHIPPED,
        ],
        order="upcoming",
    ).items

    r_statuses = sorted(CABINET_RECEIPT_VISIBLE_STATUSES)
    s_statuses = sorted(CABINET_DISPATCH_VISIBLE_STATUSES)
    r_ops = sorted(CABINET_RECEIPT_OPS_VISIBLE)
    s_ops = sorted(CABINET_DISPATCH_OPS_VISIBLE)
    events_rows = connection.execute(
        f"""
        SELECT 'receipt' AS doc_kind, o.doc_id, d.doc_number, o.op_type, o.qty, o.comment, o.created_at
        FROM receipt_ops o
        JOIN receipt_docs d ON d.id = o.doc_id
        WHERE d.client_id = ? AND d.is_deleted = 0
          AND d.status IN ({",".join("?" for _ in r_statuses)})
          AND o.op_type IN ({",".join("?" for _ in r_ops)})
        UNION ALL
        SELECT 'shipment' AS doc_kind, o.doc_id, d.doc_number, o.op_type, NULL AS qty, o.comment, o.created_at
        FROM dispatch_ops o
        JOIN dispatch_docs d ON d.id = o.doc_id
        WHERE d.client_id = ? AND d.is_deleted = 0
          AND d.status IN ({",".join("?" for _ in s_statuses)})
          AND o.op_type IN ({",".join("?" for _ in s_ops)})
        ORDER BY created_at DESC
        LIMIT 10
        """,
        [client_id, *r_statuses, *r_ops, client_id, *s_statuses, *s_ops],
    ).fetchall()

    return CabinetSummaryResponse(
        totals=cabinet_balance_totals(connection, client_id=client_id),
        active_receipts=active_receipts,
        active_shipments=active_shipments,
        events=[
            CabinetEventItem(
                doc_kind=str(r["doc_kind"]),
                doc_id=str(r["doc_id"]),
                doc_number=str(r["doc_number"]),
                op_type=str(r["op_type"]),
                qty=int(r["qty"]) if r["qty"] is not None else None,
                comment=r["comment"],
                created_at=str(r["created_at"]),
            )
            for r in events_rows
        ],
    )


def cabinet_packing_report(
    connection,
    *,
    client_id: str,
    date_from: str | None,
    date_to: str | None,
    search: str | None,
) -> CabinetPackingReportResponse:
    from modules.shipments.service import packing_productivity

    raw = packing_productivity(
        connection,
        date_from=date_from,
        date_to=date_to,
        client_id=client_id,
        search=search,
    )
    return CabinetPackingReportResponse(
        days=[
            CabinetPackingReportDay(
                packed_date=str(d["packed_date"]),
                good=int(d["good"]),
                defect=int(d["defect"]),
                total=int(d["total"]),
                sku_count=int(d["sku_count"]),
                doc_count=int(d["doc_count"]),
                rows=[
                    CabinetPackingReportRow(
                        product_sku=r["product_sku"],
                        product_name=r["product_name"],
                        good=int(r["good"]),
                        defect=int(r["defect"]),
                        total=int(r["total"]),
                    )
                    for r in d["rows"]
                ],
            )
            for d in raw["days"]
        ],
        total_good=int(raw["total_good"]),
        total_defect=int(raw["total_defect"]),
        total=int(raw["total"]),
    )


def get_cabinet_profile(connection, *, client_id: str) -> CabinetProfileResponse:
    client_row = connection.execute(
        "SELECT id, name FROM clients WHERE id = ? AND COALESCE(is_deleted, 0) = 0",
        (client_id,),
    ).fetchone()
    if not client_row:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Клиент не найден")
    stores_rows = connection.execute(
        """SELECT id, name, is_active FROM client_stores
           WHERE client_id = ? AND COALESCE(is_deleted, 0) = 0
           ORDER BY LOWER(name)""",
        (client_id,),
    ).fetchall()
    return CabinetProfileResponse(
        client={"id": str(client_row["id"]), "name": str(client_row["name"])},
        stores=[
            {"id": str(r["id"]), "name": str(r["name"]), "is_active": bool(r["is_active"])}
            for r in stores_rows
        ],
    )


# Экспорт остатков — формат согласован с клиентом (образец «Pack Men Остатки на дата»):
# «На хранении» = всё годное вне «Готов к отгрузке» (хранение + упаковка + упаковано),
# «Брак» = брак по всем зонам; сумма трёх колонок = весь остаток позиции.
BALANCES_EXPORT_LIMIT = 100_000


def build_cabinet_balances_xlsx(
    connection,
    *,
    client_id: str,
    search: str | None,
    only_positive: bool,
    has_defect: bool,
) -> tuple[bytes, str]:
    """Возвращает (содержимое xlsx, имя файла «<Клиент> Остатки на ДД.ММ.ГГГГ.xlsx»)."""
    result = get_balances(
        connection,
        page=1,
        limit=BALANCES_EXPORT_LIMIT,
        client_id=client_id,
        search=search,
        only_positive=only_positive,
        has_defect=has_defect,
    )

    wb = Workbook()
    ws = wb.active
    ws.title = "Остатки"

    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(fill_type="solid", start_color="305496")
    thin = Side(style="thin", color="B7C1D1")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal="center", vertical="center")

    headers = ["Наименование", "Артикул", "На хранении", "Готов к отгрузке", "Брак"]
    for col, title in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=title)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border
        cell.alignment = center
    ws.row_dimensions[1].height = 22
    ws.freeze_panes = "A2"

    widths = {"A": 45, "B": 20, "C": 14, "D": 16, "E": 12}
    for letter, width in widths.items():
        ws.column_dimensions[letter].width = width

    for row_idx, item in enumerate(result.items, start=2):
        name = ", ".join(p for p in (item.product_name, item.color_name, item.size_name) if p)
        in_storage = item.storage_good + item.packing_good + item.packed_good
        defect = item.storage_defect + item.packing_defect + item.packed_defect + item.ready_defect
        values = [name, item.product_sku, in_storage, item.ready_good, defect]
        for col, value in enumerate(values, start=1):
            cell = ws.cell(row=row_idx, column=col, value=value)
            cell.border = border
            if col >= 3:
                cell.number_format = "#,##0"

    if result.items:
        ws.auto_filter.ref = f"A1:E{len(result.items) + 1}"

    client_row = connection.execute(
        "SELECT name FROM clients WHERE id = ?", (client_id,)
    ).fetchone()
    client_name = str(client_row["name"]).strip() if client_row and client_row["name"] else ""
    date_str = business_today().strftime("%d.%m.%Y")
    filename = f"{client_name + ' ' if client_name else ''}Остатки на {date_str}.xlsx"

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue(), filename
