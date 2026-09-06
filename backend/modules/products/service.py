from __future__ import annotations

import json
import re
from typing import Any, Mapping, NamedTuple
from uuid import uuid4

from fastapi import HTTPException, status
from config import (
    MAX_UPLOAD_BYTES,
    PRODUCT_IMPORT_ACTION_APPEND,
    PRODUCT_IMPORT_ACTION_CREATE,
    PRODUCT_IMPORT_ACTION_ERROR,
    PRODUCT_IMPORT_ACTION_LABELS,
    PRODUCT_IMPORT_ACTION_SKIP,
    PRODUCT_IMPORT_BATCH_TTL_HOURS,
    PRODUCT_IMPORT_MAX_BYTES,
    PRODUCT_IMPORT_MAX_ROWS,
    PRODUCT_IMPORT_STATUS_PREVIEW,
    PRODUCT_LIST_SORT_COLUMNS,
    UPLOADS_DIR,
)
from dbconn import escape_like, get_connection

from .schemas import (
    BarcodeLabelItem,
    BarcodeLabelMissingItem,
    BarcodeLabelRequestItem,
    BarcodeLabelsResponse,
    ProductCreateDimensionBlock,
    ProductCreateMeta,
    ProductItem,
    ProductListResponse,
    ProductUpdateRequest,
    ProductVariantDimension,
    ProductVariantFindItem,
    ProductVariantFindResponse,
    ProductVariantItem,
    ProductVariantsPatchRequest,
    MessageResponse,
)
from utils import barcode_svg as _barcode_svg, now_iso as _now



def _normalize_name(name: str) -> str:
    normalized = name.strip()
    if not normalized:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Поле Название обязательно")
    return normalized


def _normalize_sku(sku: str) -> str:
    normalized = sku.strip()
    if not normalized:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Поле Штрих-код товара обязательно")
    return normalized


def _fold_ci_str(x: object) -> str:
    if x is None:
        return ""
    return str(x).lower().replace("ё", "е")


def _ci_substring_like_param(raw: str) -> str:
    return f"%{escape_like(_fold_ci_str(str(raw).strip()))}%"


def _order_sql_from_sort_param(sort: str | None, allowed: dict[str, str]) -> str | None:
    if not sort or not str(sort).strip():
        return None
    head, sep, tail = str(sort).strip().rpartition("_")
    if not sep or tail.lower() not in ("asc", "desc"):
        return None
    if head not in allowed:
        return None
    return f"{allowed[head]} {tail.upper()}"


def _decode_images_json(raw: str | None) -> list[str]:
    if not raw:
        return []
    try:
        v = json.loads(raw)
        return [str(x) for x in v] if isinstance(v, list) else []
    except json.JSONDecodeError:
        return []


def _encode_images_json(urls: list[str]) -> str:
    return json.dumps(urls, ensure_ascii=False)


def _product_card_image_urls(gallery_json_val: str | None, image_url_val: str | None) -> list[str]:
    if gallery_json_val and str(gallery_json_val).strip():
        try:
            v = json.loads(str(gallery_json_val))
            if isinstance(v, list) and v:
                return [str(x).strip() for x in v if str(x).strip()]
        except json.JSONDecodeError:
            pass
    if image_url_val and str(image_url_val).strip():
        return [str(image_url_val).strip()]
    return []


def _product_image_extension(content_type: str | None, original_filename: str | None) -> str | None:
    if not content_type and not original_filename:
        return None
    if content_type:
        ct = content_type.split(";", 1)[0].strip().lower()
        if ct in ("image/jpeg", "image/jpg"):
            return ".jpg"
        if ct == "image/png":
            return ".png"
        if ct in ("image/heic", "image/heif"):
            return ".heic"
    if original_filename:
        parts = original_filename.rsplit(".", 1)
        if len(parts) == 2 and parts[1].lower() in ("heic", "heif"):
            return f".{parts[1].lower()}"
    return None


def _sku_token_from_label(name: str) -> str:
    s = re.sub(r"\s+", "-", str(name).strip().upper())
    s = re.sub(r"[^0-9A-ZА-ЯЁ\-]", "", s)
    s = re.sub(r"-+", "-", s).strip("-") or "X"
    return s[:18]


def _color_size_labels_for_skus(
    connection: Any, color_ids: set[str], size_ids: set[str]
) -> tuple[dict[str, str], dict[str, str]]:
    colors: dict[str, str] = {}
    if color_ids:
        placeholders = ",".join("?" * len(color_ids))
        for r in connection.execute(
            f"SELECT id, name FROM colors WHERE id IN ({placeholders})",
            list(color_ids),
        ).fetchall():
            colors[str(r["id"])] = str(r["name"])
    sizes: dict[str, str] = {}
    if size_ids:
        placeholders = ",".join("?" * len(size_ids))
        for r in connection.execute(
            f"SELECT id, name FROM sizes WHERE id IN ({placeholders})",
            list(size_ids),
        ).fetchall():
            sizes[str(r["id"])] = str(r["name"])
    return colors, sizes


def _variant_identity_key(
    sku_base: str, color_id: str, size_id: str | None, *, requires_size: bool
) -> tuple[str, str, str]:
    base = str(sku_base).strip()
    cid = str(color_id).strip().lower()
    sid = str(size_id).strip().lower() if size_id and requires_size else ""
    return (base, cid, sid)


def _norm_variant_row_id(value: str | None) -> str | None:
    if value is None:
        return None
    s = str(value).strip()
    return s.lower() if s else None


def _variant_sku_in_use(
    connection: Any,
    sku: str,
    exclude_variant_id: str | None,
    client_id: str | None,
) -> bool:
    if not str(sku).strip():
        return False
    q = """
        SELECT 1
        FROM product_variants v
        JOIN products p ON p.id = v.product_id
        WHERE v.sku = ?
          AND COALESCE(v.is_deleted, 0) = 0
          AND COALESCE(p.is_deleted, 0) = 0
    """
    params: list[object] = [sku]
    if client_id is not None:
        q += " AND p.client_id = ?"
        params.append(client_id)
    if exclude_variant_id:
        q += " AND v.id != ?"
        params.append(exclude_variant_id)
    return connection.execute(q, tuple(params)).fetchone() is not None


def _sku_taken_for_client_except_product(
    connection: Any,
    sku: str,
    client_id: str,
    exclude_product_id: str | None,
) -> bool:
    if not str(sku).strip():
        return False
    product_sql = """
        SELECT 1
        FROM products
        WHERE sku = ?
          AND client_id = ?
          AND COALESCE(is_deleted, 0) = 0
    """
    product_params: list[object] = [sku, client_id]
    if exclude_product_id:
        product_sql += " AND id != ?"
        product_params.append(exclude_product_id)
    if connection.execute(product_sql, tuple(product_params)).fetchone():
        return True

    variant_sql = """
        SELECT 1
        FROM product_variants v
        JOIN products p ON p.id = v.product_id
        WHERE v.sku = ?
          AND p.client_id = ?
          AND COALESCE(v.is_deleted, 0) = 0
          AND COALESCE(p.is_deleted, 0) = 0
    """
    variant_params: list[object] = [sku, client_id]
    if exclude_product_id:
        variant_sql += " AND v.product_id != ?"
        variant_params.append(exclude_product_id)
    return connection.execute(variant_sql, tuple(variant_params)).fetchone() is not None


def _generate_variant_sku_for_patch(
    connection: Any,
    *,
    sku_base: str,
    color_id: str,
    size_id: str | None,
    exclude_variant_id: str | None,
    client_id: str | None,
) -> str:
    base_norm = _normalize_sku(sku_base)
    color_labels, size_labels = _color_size_labels_for_skus(
        connection, {color_id}, {size_id} if size_id else set()
    )
    ct = _sku_token_from_label(color_labels.get(color_id, "C"))
    if size_id:
        st = _sku_token_from_label(size_labels.get(size_id, "S"))
        candidate = f"{base_norm}-{ct}-{st}"
    else:
        candidate = f"{base_norm}-{ct}"
    sku = candidate
    salt = 0
    while True:
        q = """
            SELECT v.id
            FROM product_variants v
            JOIN products p ON p.id = v.product_id
            WHERE v.sku = ?
              AND COALESCE(v.is_deleted, 0) = 0
              AND COALESCE(p.is_deleted, 0) = 0
        """
        params: list[object] = [sku]
        if client_id is not None:
            q += " AND p.client_id = ?"
            params.append(client_id)
        if exclude_variant_id:
            q += " AND v.id != ?"
            params.append(exclude_variant_id)
        if not connection.execute(q, tuple(params)).fetchone():
            return sku
        salt += 1
        sku = f"{candidate}-{salt}"


def _sku_dim_token(length: float, width: float, height: float) -> str:
    a, b, c = int(round(float(length))), int(round(float(width))), int(round(float(height)))
    return f"{a}X{b}X{c}"


def _require_active_product_type(connection: Any, type_id: str) -> str:
    tid = type_id.strip()
    row = connection.execute(
        "SELECT id FROM product_types WHERE id = ? AND is_active = 1 AND COALESCE(is_deleted, 0) = 0",
        (tid,),
    ).fetchone()
    if not row:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Тип товара: недопустимое или неактивное значение")
    return tid


def _require_active_client(connection: Any, raw: str | None) -> str:
    if raw is None or str(raw).strip() == "":
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Поле Клиент обязательно")
    cid = str(raw).strip()
    row = connection.execute(
        "SELECT id FROM clients WHERE id = ? AND is_active = 1 AND COALESCE(is_deleted, 0) = 0",
        (cid,),
    ).fetchone()
    if not row:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Клиент: недопустимое или неактивное значение")
    return cid


def _optional_active_client(connection: Any, raw: str | None) -> str | None:
    if raw is None or str(raw).strip() == "":
        return None
    cid = str(raw).strip()
    row = connection.execute(
        "SELECT id FROM clients WHERE id = ? AND is_active = 1 AND COALESCE(is_deleted, 0) = 0",
        (cid,),
    ).fetchone()
    if not row:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Клиент: недопустимое или неактивное значение")
    return cid


def _require_active_color_id(connection: Any, cid: str) -> str:
    rid = cid.strip()
    row = connection.execute(
        "SELECT id FROM colors WHERE id = ? AND is_active = 1 AND COALESCE(is_deleted, 0) = 0",
        (rid,),
    ).fetchone()
    if not row:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Цвет: недопустимое или неактивное значение")
    return rid


def _require_active_size_id(connection: Any, sid: str) -> str:
    rid = sid.strip()
    row = connection.execute(
        "SELECT id FROM sizes WHERE id = ? AND is_active = 1 AND COALESCE(is_deleted, 0) = 0",
        (rid,),
    ).fetchone()
    if not row:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Размер: недопустимое или неактивное значение")
    return rid


def _product_type_flags(connection: Any, type_id: str) -> tuple[bool, bool]:
    row = connection.execute(
        "SELECT requires_color, requires_size FROM product_types WHERE id = ?",
        (type_id,),
    ).fetchone()
    if not row:
        return False, False
    return bool(row["requires_color"]), bool(row["requires_size"])


def _rebase_variant_skus_for_new_product_base(
    connection: Any,
    *,
    product_id: str,
    old_base_sku: str,
    new_base_sku: str,
    updated_at: str,
    client_id: str | None,
) -> None:
    old_b = old_base_sku.strip()
    new_b = new_base_sku.strip()
    rows = connection.execute(
        "SELECT id, sku, color_id, size_id FROM product_variants WHERE product_id = ? AND COALESCE(is_deleted, 0) = 0 ORDER BY LOWER(sku) ASC",
        (product_id,),
    ).fetchall()
    computed: list[tuple[str, str, str, str | None]] = []
    for r in rows:
        vid = str(r["id"])
        cur = str(r["sku"]).strip()
        color_id = str(r["color_id"])
        sz_id = str(r["size_id"]) if r["size_id"] else None
        if cur == old_b or cur.startswith(old_b + "-"):
            prop = new_b + cur[len(old_b):] if cur.startswith(old_b + "-") else new_b
        else:
            prop = _generate_variant_sku_for_patch(connection, sku_base=new_b, color_id=color_id, size_id=sz_id, exclude_variant_id=vid, client_id=client_id)
        if prop == new_b:
            prop = _generate_variant_sku_for_patch(connection, sku_base=new_b, color_id=color_id, size_id=sz_id, exclude_variant_id=vid, client_id=client_id)
        computed.append((vid, prop, color_id, sz_id))

    used_local: set[str] = set()
    finals: list[tuple[str, str]] = []
    for vid, prop, color_id, sz_id in computed:
        p = prop
        n = 0
        while n < 500:
            n += 1
            if p not in used_local and not _variant_sku_in_use(connection, p, vid, client_id):
                break
            p = _generate_variant_sku_for_patch(connection, sku_base=new_b, color_id=color_id, size_id=sz_id, exclude_variant_id=vid, client_id=client_id)
        else:
            raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Не удалось подобрать уникальные штрих-коды вариантов")
        used_local.add(p)
        finals.append((vid, p))

    for vid, p in finals:
        connection.execute("UPDATE product_variants SET sku = ?, updated_at = ? WHERE id = ?", (p, updated_at, vid))


def _assign_variant_skus_from_base(
    connection: Any,
    *,
    product_id: str,
    new_base_sku: str,
    updated_at: str,
    client_id: str | None,
) -> None:
    """Присвоить SKU вариантам товара, ранее «ожидавшего SKU».

    Генерирует SKU из нового базового по цвету/размеру каждого варианта и снимает
    признак ожидания. Вызывается в момент, когда у pending-товара появился базовый SKU.
    """
    new_b = new_base_sku.strip()
    rows = connection.execute(
        "SELECT id, color_id, size_id FROM product_variants WHERE product_id = ? AND COALESCE(is_deleted, 0) = 0 ORDER BY created_at ASC",
        (product_id,),
    ).fetchall()
    used_local: set[str] = set()
    for r in rows:
        vid = str(r["id"])
        color_id = str(r["color_id"]) if r["color_id"] else ""
        sz_id = str(r["size_id"]) if r["size_id"] else None
        sku = _generate_variant_sku_for_patch(
            connection, sku_base=new_b, color_id=color_id, size_id=sz_id, exclude_variant_id=vid, client_id=client_id
        )
        n = 0
        while sku in used_local and n < 500:
            n += 1
            sku = f"{sku}-{n}"
        used_local.add(sku)
        connection.execute(
            "UPDATE product_variants SET sku = ?, sku_pending = 0, updated_at = ? WHERE id = ?",
            (sku, updated_at, vid),
        )


def _fill_empty_line_sku_snapshots(connection: Any, *, product_id: str, sku: str) -> None:
    for table_name in ("shipment_lines", "dispatch_lines"):
        connection.execute(
            f"""
            UPDATE {table_name}
            SET product_sku = ?
            WHERE product_id = ?
              AND COALESCE(NULLIF(TRIM(product_sku), ''), '') = ''
            """,
            (sku, product_id),
        )


def assign_product_sku_if_missing(
    connection: Any,
    *,
    product_id: str,
    sku_base: str | None,
    updated_at: str,
    user_id: str | None = None,
) -> str:
    sku_t = str(sku_base or "").strip()
    if not sku_t:
        return ""
    row = connection.execute(
        """
        SELECT sku, COALESCE(sku_pending, 0) AS sku_pending, client_id
        FROM products
        WHERE id = ? AND COALESCE(is_deleted, 0) = 0
        """,
        (product_id,),
    ).fetchone()
    if not row:
        return sku_t
    cur_sku = str(row["sku"] or "").strip()
    if cur_sku:
        return cur_sku

    new_sku = _normalize_sku(sku_t)
    client_id = str(row["client_id"]) if row["client_id"] else None
    if client_id is not None and _sku_taken_for_client_except_product(connection, new_sku, client_id, product_id):
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Базовый штрих-код уже занят у этого клиента")

    _assign_variant_skus_from_base(
        connection,
        product_id=product_id,
        new_base_sku=new_sku,
        updated_at=updated_at,
        client_id=client_id,
    )
    if user_id:
        connection.execute(
            "UPDATE products SET sku = ?, sku_pending = 0, updated_at = ?, updated_by_id = ? WHERE id = ?",
            (new_sku, updated_at, user_id, product_id),
        )
    else:
        connection.execute(
            "UPDATE products SET sku = ?, sku_pending = 0, updated_at = ? WHERE id = ?",
            (new_sku, updated_at, product_id),
        )
    _fill_empty_line_sku_snapshots(connection, product_id=product_id, sku=new_sku)
    return new_sku


def _soft_delete_variants_for_product(connection: Any, product_id: str, admin_id: str, ts: str) -> None:
    connection.execute(
        """
        UPDATE product_variants
        SET is_deleted = 1, deleted_at = ?, deleted_by_id = ?, updated_at = ?, is_active = 0
        WHERE product_id = ? AND COALESCE(is_deleted, 0) = 0
        """,
        (ts, admin_id, ts, product_id),
    )


def _build_variant_rows_for_create(
    connection: Any,
    *,
    requires_size: bool,
    sku_base: str,
    client_id: str,
    color_ids: list[str],
    dimensions: list[ProductCreateDimensionBlock],
    sku_pending: bool = False,
) -> list[dict]:
    seen_keys: set[tuple] = set()
    out: list[dict] = []
    used_skus: set[str] = set()
    size_ids_collect: set[str] = set()
    for block in dimensions:
        for s in block.sizes:
            size_ids_collect.add(s)

    if not dimensions:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Укажите хотя бы один блок габаритов")

    color_labels, size_labels = _color_size_labels_for_skus(connection, set(color_ids), size_ids_collect)

    if requires_size:
        for block in dimensions:
            for cid in color_ids:
                for szid in block.sizes:
                    key = (cid, szid)
                    if key in seen_keys:
                        continue
                    seen_keys.add(key)
                    if sku_pending:
                        sku = ""
                    else:
                        ct = _sku_token_from_label(color_labels.get(cid, "C"))
                        st = _sku_token_from_label(size_labels.get(szid, "S"))
                        sku = f"{sku_base}-{ct}-{st}"
                        n = 0
                        base_sku = sku
                        while sku in used_skus or _variant_sku_in_use(connection, sku, None, client_id):
                            n += 1
                            sku = f"{base_sku}-{n}"
                        used_skus.add(sku)
                    out.append({
                        "color_id": cid, "size_id": szid,
                        "length": float(block.length), "width": float(block.width), "height": float(block.height),
                        "sku": sku, "images_json": "[]",
                    })
    else:
        multi_block = len(dimensions) > 1
        for block in dimensions:
            for cid in color_ids:
                key = (cid, round(block.length, 4), round(block.width, 4), round(block.height, 4))
                if key in seen_keys:
                    continue
                seen_keys.add(key)
                if sku_pending:
                    sku = ""
                else:
                    ct = _sku_token_from_label(color_labels.get(cid, "C"))
                    if multi_block:
                        dim_tok = _sku_dim_token(block.length, block.width, block.height)
                        sku = f"{sku_base}-{ct}-{dim_tok}"
                    else:
                        sku = f"{sku_base}-{ct}"
                    n = 0
                    base_sku = sku
                    while sku in used_skus or _variant_sku_in_use(connection, sku, None, client_id):
                        n += 1
                        sku = f"{base_sku}-{n}"
                    used_skus.add(sku)
                out.append({
                    "color_id": cid, "size_id": None,
                    "length": float(block.length), "width": float(block.width), "height": float(block.height),
                    "sku": sku, "images_json": "[]",
                })
    return out


def _sync_product_variants_from_request(
    connection: Any,
    product_id: str,
    payload: ProductVariantsPatchRequest,
    admin_id: str,
) -> None:
    prow = connection.execute(
        """
        SELECT p.sku AS sku_base, COALESCE(p.sku_pending, 0) AS sku_pending,
               p.client_id, COALESCE(pt.requires_size, 0) AS requires_size,
               COALESCE(p.is_deleted, 0) AS is_deleted
        FROM products p
        JOIN product_types pt ON pt.id = p.type_id
        WHERE p.id = ?
        """,
        (product_id,),
    ).fetchone()
    if not prow:
        raise HTTPException(status_code=404, detail="Товар не найден")
    if bool(prow["is_deleted"]):
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Товар удалён. Восстановите его перед редактированием вариантов.")
    sku_base = str(prow["sku_base"])
    prod_pending = bool(prow["sku_pending"])
    client_id = str(prow["client_id"]) if prow["client_id"] else None
    requires_size = bool(prow["requires_size"])
    now = _now()

    existing_rows = connection.execute(
        "SELECT id FROM product_variants WHERE product_id = ?",
        (product_id,),
    ).fetchall()
    existing_by_norm: dict[str, str] = {}
    for r in existing_rows:
        raw_id = str(r["id"])
        n = _norm_variant_row_id(raw_id)
        if n:
            existing_by_norm[n] = raw_id
    incoming_norm = {_norm_variant_row_id(str(v.id)) for v in payload.variants if v.id}
    incoming_norm.discard(None)
    for norm_id, raw_id in existing_by_norm.items():
        if norm_id not in incoming_norm:
            connection.execute(
                """
                UPDATE product_variants
                SET is_deleted = 1, deleted_at = ?, deleted_by_id = ?, updated_at = ?, is_active = 0
                WHERE id = ? AND product_id = ? AND COALESCE(is_deleted, 0) = 0
                """,
                (now, admin_id, now, raw_id, product_id),
            )

    dup_detail = (
        "Дублируется сочетание штрих-кода товара, цвета и размера"
        if requires_size
        else "Дублируется сочетание штрих-кода товара и цвета"
    )
    rows_to_apply: list[tuple] = []
    identity_keys: list[tuple[str, str, str]] = []
    for item in payload.variants:
        eff_size = item.size_id if requires_size else None
        if requires_size and (eff_size is None or str(eff_size).strip() == ""):
            raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Для этого типа товара укажите размер варианта")
        if item.color_id:
            _require_active_color_id(connection, item.color_id)
        if eff_size:
            _require_active_size_id(connection, eff_size)
        identity_keys.append(_variant_identity_key(sku_base, item.color_id or "", eff_size, requires_size=requires_size))
        rows_to_apply.append((item, eff_size))

    if len(identity_keys) != len(set(identity_keys)):
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail=dup_detail)

    for item, eff_size in rows_to_apply:
        imgs = _encode_images_json(item.images)
        if item.id:
            rid = str(item.id)
            own = connection.execute(
                "SELECT id, color_id, size_id FROM product_variants WHERE id = ? AND product_id = ?",
                (rid, product_id),
            ).fetchone()
            if not own:
                raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Неизвестный вариант")
            has_receipts = connection.execute(
                """
                SELECT 1 FROM receipt_lines rl
                JOIN receipt_docs rd ON rd.id = rl.doc_id
                WHERE rl.product_id = ?
                  AND rl.color_id IS NOT DISTINCT FROM ?
                  AND rl.size_id IS NOT DISTINCT FROM ?
                  AND rl.is_deleted = 0 AND rd.is_deleted = 0
                LIMIT 1
                """,
                (product_id, own["color_id"], own["size_id"]),
            ).fetchone()
            if has_receipts:
                if item.color_id != own["color_id"] or eff_size != own["size_id"]:
                    raise HTTPException(
                        status_code=status.HTTP_409_CONFLICT,
                        detail="Нельзя изменить цвет или размер варианта: по нему зафиксированы поступления",
                    )
            if prod_pending:
                new_sku = ""
            else:
                new_sku = (
                    _normalize_sku(item.sku)
                    if item.sku and str(item.sku).strip()
                    else _generate_variant_sku_for_patch(connection, sku_base=sku_base, color_id=item.color_id, size_id=eff_size, exclude_variant_id=rid, client_id=client_id)
                )
                if _variant_sku_in_use(connection, new_sku, rid, client_id):
                    raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="SKU варианта уже занят у этого клиента")
            connection.execute(
                """
                UPDATE product_variants
                SET color_id = ?, size_id = ?, length = ?, width = ?, height = ?,
                    sku = ?, sku_pending = ?, images_json = ?, is_active = ?, updated_at = ?, client_id = ?,
                    is_deleted = 0, deleted_at = NULL, deleted_by_id = NULL
                WHERE id = ? AND product_id = ?
                """,
                (item.color_id, eff_size, float(item.dimension.length), float(item.dimension.width), float(item.dimension.height), new_sku, 1 if prod_pending else 0, imgs, 1 if item.is_active else 0, now, client_id, rid, product_id),
            )
        else:
            if prod_pending:
                new_sku = ""
            else:
                new_sku = (
                    _normalize_sku(item.sku)
                    if item.sku and str(item.sku).strip()
                    else _generate_variant_sku_for_patch(connection, sku_base=sku_base, color_id=item.color_id, size_id=eff_size, exclude_variant_id=None, client_id=client_id)
                )
                if _variant_sku_in_use(connection, new_sku, None, client_id):
                    raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="SKU варианта уже занят у этого клиента")
            connection.execute(
                """
                INSERT INTO product_variants (id, product_id, client_id, color_id, size_id, length, width, height, sku, sku_pending, images_json, is_active, created_at, is_deleted)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, 0)
                """,
                (str(uuid4()), product_id, client_id, item.color_id, eff_size, float(item.dimension.length), float(item.dimension.width), float(item.dimension.height), new_sku, 1 if prod_pending else 0, imgs, 1 if item.is_active else 0, now),
            )


def change_variant_identity(
    connection: Any,
    product_id: str,
    variant_id: str,
    *,
    color_id: str | None,
    size_id: str | None,
    sku: str | None,
    admin_id: str,
) -> MessageResponse:
    """Смена цвета/размера варианта с переносом истории на новый ключ.

    Идентичность остатка — тройка (product_id, color_id, size_id), поэтому обычный
    PATCH вариантов запрещает менять оси при наличии поступлений (иначе журнальный
    остаток «осиротеет»). Эта операция решает случай «цвет узнали позже»: атомарно
    пере-ключевывает ВСЕ носители ключа (журнал zone_relocations, строки поступлений,
    задач упаковки и отгрузок) — остатки, резервы и гейты продолжают сходиться.
    Снимки имён (color_name и т.п.) в журнале и строках не трогаем: документы
    показывают то, что было записано на момент операции. След — в append-only
    variant_identity_changes.
    """
    prow = connection.execute(
        """
        SELECT p.sku AS sku_base, COALESCE(p.sku_pending, 0) AS sku_pending,
               p.client_id, COALESCE(pt.requires_size, 0) AS requires_size,
               COALESCE(p.is_deleted, 0) AS is_deleted
        FROM products p
        JOIN product_types pt ON pt.id = p.type_id
        WHERE p.id = ?
        """,
        (product_id,),
    ).fetchone()
    if not prow:
        raise HTTPException(status_code=404, detail="Товар не найден")
    if bool(prow["is_deleted"]):
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Товар удалён. Восстановите его перед изменением вариантов.")
    requires_size = bool(prow["requires_size"])
    prod_pending = bool(prow["sku_pending"])
    client_id = str(prow["client_id"]) if prow["client_id"] else None

    variant = connection.execute(
        "SELECT id, color_id, size_id, sku FROM product_variants "
        "WHERE id = ? AND product_id = ? AND COALESCE(is_deleted, 0) = 0",
        (variant_id, product_id),
    ).fetchone()
    if not variant:
        raise HTTPException(status_code=404, detail="Вариант не найден")

    new_color = _require_active_color_id(connection, color_id) if color_id and str(color_id).strip() else None
    eff_size = size_id if requires_size else None
    if requires_size and (eff_size is None or str(eff_size).strip() == ""):
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Для этого типа товара укажите размер варианта")
    new_size = _require_active_size_id(connection, eff_size) if eff_size else None

    old_color = variant["color_id"]
    old_size = variant["size_id"]
    old_sku = str(variant["sku"] or "")
    if new_color == old_color and new_size == old_size:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Цвет и размер варианта не изменились")

    clash = connection.execute(
        "SELECT 1 FROM product_variants WHERE product_id = ? AND id != ? "
        "AND COALESCE(is_deleted, 0) = 0 "
        "AND color_id IS NOT DISTINCT FROM ? AND size_id IS NOT DISTINCT FROM ?",
        (product_id, variant_id, new_color, new_size),
    ).fetchone()
    if clash:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="У товара уже есть вариант с таким сочетанием цвета и размера",
        )

    if prod_pending:
        new_sku = ""
    elif sku and str(sku).strip():
        new_sku = _normalize_sku(sku)
        if _variant_sku_in_use(connection, new_sku, variant_id, client_id):
            raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="SKU варианта уже занят у этого клиента")
    elif new_color:
        new_sku = _generate_variant_sku_for_patch(
            connection, sku_base=str(prow["sku_base"]), color_id=new_color,
            size_id=new_size, exclude_variant_id=variant_id, client_id=client_id,
        )
    else:
        new_sku = old_sku

    key_where = "product_id = ? AND color_id IS NOT DISTINCT FROM ? AND size_id IS NOT DISTINCT FROM ?"
    key_params = (product_id, old_color, old_size)
    moved = {}
    for name, table in (
        ("journal_rows", "zone_relocations"),
        ("receipt_rows", "receipt_lines"),
        ("shipment_rows", "shipment_lines"),
        ("dispatch_rows", "dispatch_lines"),
    ):
        cur = connection.execute(
            f"UPDATE {table} SET color_id = ?, size_id = ? WHERE {key_where}",
            (new_color, new_size, *key_params),
        )
        moved[name] = int(cur.rowcount or 0)

    now = _now()
    connection.execute(
        "UPDATE product_variants SET color_id = ?, size_id = ?, sku = ?, updated_at = ? "
        "WHERE id = ? AND product_id = ?",
        (new_color, new_size, new_sku, now, variant_id, product_id),
    )
    connection.execute(
        "UPDATE products SET updated_at = ?, updated_by_id = ? WHERE id = ?",
        (now, admin_id, product_id),
    )

    color_labels, size_labels = _color_size_labels_for_skus(
        connection,
        {c for c in (old_color, new_color) if c},
        {s for s in (old_size, new_size) if s},
    )
    connection.execute(
        """INSERT INTO variant_identity_changes
           (id, product_id, variant_id,
            old_color_id, old_color_name, new_color_id, new_color_name,
            old_size_id, old_size_name, new_size_id, new_size_name,
            old_sku, new_sku,
            journal_rows, receipt_rows, shipment_rows, dispatch_rows,
            created_at, created_by)
           VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)""",
        (str(uuid4()), product_id, variant_id,
         old_color, color_labels.get(old_color or ""), new_color, color_labels.get(new_color or ""),
         old_size, size_labels.get(old_size or ""), new_size, size_labels.get(new_size or ""),
         old_sku, new_sku,
         moved["journal_rows"], moved["receipt_rows"], moved["shipment_rows"], moved["dispatch_rows"],
         now, admin_id),
    )
    return MessageResponse(message="Цвет/размер варианта изменены, остатки и история перенесены")


def _find_variant_row_for_receipt(
    connection: Any,
    sku_norm: str,
    color_id: str,
    size_id: str | None,
) -> tuple[Mapping[str, Any] | None, bool]:
    rows = connection.execute(
        """
        SELECT v.id AS variant_id, v.product_id, v.sku AS variant_sku, p.sku AS product_sku,
               v.color_id, v.size_id, v.length, v.width, v.height,
               p.name AS product_name, p.gallery_json, p.image_url,
               COALESCE(pt.requires_size, 0) AS requires_size,
               pt.name AS product_type_name,
               cl.name AS client_name
        FROM product_variants v
        JOIN products p ON p.id = v.product_id
        JOIN product_types pt ON pt.id = p.type_id
        LEFT JOIN clients cl ON cl.id = p.client_id
        WHERE COALESCE(v.is_deleted, 0) = 0 AND COALESCE(p.is_deleted, 0) = 0
          AND COALESCE(v.is_active, 1) = 1 AND COALESCE(p.is_active, 1) = 1
          AND v.color_id = ?
          AND (
            LOWER(TRIM(COALESCE(v.sku, ''))) = LOWER(?)
            OR LOWER(TRIM(COALESCE(p.sku, ''))) = LOWER(?)
          )
        """,
        (color_id, sku_norm, sku_norm),
    ).fetchall()
    if not rows:
        return None, False
    req_size = bool(rows[0]["requires_size"])
    if not req_size:
        if len(rows) != 1:
            return None, False
        return rows[0], False
    user_sz = (size_id or "").strip()
    if len(rows) == 1:
        r = rows[0]
        vs = str(r["size_id"] or "")
        if not user_sz:
            return None, True
        if vs != user_sz:
            return None, True
        return r, False
    if not user_sz:
        return None, True
    matches = [r for r in rows if str(r["size_id"] or "") == user_sz]
    if len(matches) != 1:
        return None, True
    return matches[0], False


def _resolve_actuality_filter(connection: Any, actuality_id: str | None) -> bool | None:
    if actuality_id is None:
        return None
    aid = str(actuality_id).strip()
    if not aid:
        return None
    row = connection.execute(
        "SELECT maps_is_active FROM record_actuality WHERE id = ?",
        (aid,),
    ).fetchone()
    if not row:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Недопустимое значение фильтра актуальности")
    return bool(row["maps_is_active"])


def _row_to_product_item(row: Mapping[str, Any]) -> ProductItem:
    return ProductItem(
        id=row["id"],
        name=row["name"],
        type_id=row["type_id"],
        type_name=row["type_name"],
        sku_base=row["sku_base"],
        sku_pending=bool(row.get("sku_pending") or 0),
        weight_grams=int(row["weight_grams"]) if row["weight_grams"] is not None else None,
        items_per_box=int(row["items_per_box"]) if row.get("items_per_box") is not None else None,
        boxes_per_pallet=int(row["boxes_per_pallet"]) if row.get("boxes_per_pallet") is not None else None,
        requires_color=bool(row["requires_color"]),
        requires_size=bool(row["requires_size"]),
        client_id=row["client_id"],
        client_name=row["client_name"],
        variant_count=int(row["variant_count"] or 0),
        stock_total=max(0, int(row.get("stock_total") or 0)),
        defect_total=max(0, int(row.get("defect_total") or 0)),
        is_active=bool(row["is_active"]),
        is_deleted=bool(row["is_deleted"]),
        deleted_at=row["deleted_at"],
        deleted_by=row["deleted_by"],
        image_urls=_product_card_image_urls(row["gallery_json"], row["image_url"]),
        created_at=row["created_at"],
        created_by=row["created_by"],
        updated_at=row["updated_at"],
        updated_by=row["updated_by"],
    )


def _find_barcode_owner(connection, code: str):
    return connection.execute(
        """
        SELECT p.name AS product_name, col.name AS color_name, sz.name AS size_name
        FROM product_barcodes pb
        JOIN products p ON p.id = pb.product_id
        LEFT JOIN product_variants v ON v.id = pb.variant_id
        LEFT JOIN colors col ON col.id = v.color_id
        LEFT JOIN sizes sz ON sz.id = v.size_id
        WHERE pb.barcode = ? AND COALESCE(pb.is_deleted, 0) = 0
        LIMIT 1
        """,
        (code,),
    ).fetchone()


def _barcode_owner_label(owner) -> str:
    label = " · ".join(x for x in (owner["color_name"], owner["size_name"]) if x)
    return f"«{owner['product_name']}»" + (f" ({label})" if label else "")


# ---------------------------------------------------------------------------
# Массовая загрузка товаров из Excel
# ---------------------------------------------------------------------------

# Строка файла = ВАРИАНТ товара: варианты несут собственные габариты и штрих-коды,
# одной «плоской» строкой на товар их не описать. Товары собираются группировкой
# по базовому SKU (для позиций без SKU — по названию и типу внутри одного файла).
_IMPORT_COLUMNS: tuple[tuple[str, str, tuple[str, ...]], ...] = (
    ("sku", "SKU базовый", ("sku", "артикул", "шк", "штрих-код товара")),
    ("name", "Название товара", ("название", "name")),
    ("type", "Тип товара", ("тип", "type")),
    ("color", "Цвет", ("color",)),
    ("size", "Размер", ("size",)),
    ("length", "Длина, см", ("длина",)),
    ("width", "Ширина, см", ("ширина",)),
    ("height", "Высота, см", ("высота",)),
    ("weight", "Вес, гр", ("вес", "вес, г", "вес, гр.")),
    ("items_per_box", "Кол-во в коробе", ("количество в коробе", "в коробе")),
    ("boxes_per_pallet", "Коробов на палете", ("коробов на паллете", "на палете")),
    ("barcodes", "Штрих-код(ы)", ("штрих-коды", "штрихкод", "штрих-код варианта")),
    ("price_good", "Цена упаковки годный, ₽", ("цена упаковки годный",)),
    ("price_defect", "Цена упаковки брак, ₽", ("цена упаковки брак",)),
    ("active", "Активен", ("активность",)),
)

_IMPORT_REQUIRED_COLUMNS = ("sku", "name", "type")

_IMPORT_BARCODE_SOURCE = "Загрузка из Excel"

_IMPORT_STRUCTURE_ERROR = "Неверная структура файла"
_IMPORT_UNREADABLE_ERROR = "Не удалось прочитать файл"


class _ImportRowError(Exception):
    """Ошибка поля строки — попадает в отчёт, а не роняет разбор файла."""


def _import_collapse_ws(raw: str) -> str:
    """Трим и схлопывание внутренних пробелов: файл клиента полон случайных пробелов."""
    return re.sub(r"\s+", " ", str(raw)).strip()


def _import_cell_str(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, bool):
        return "да" if value else "нет"
    if isinstance(value, float) and float(value).is_integer():
        # Числовая ячейка со штрих-кодом иначе приедет как «4600000000001.0».
        return str(int(value))
    return _import_collapse_ws(str(value))


def _import_field(errors: list[str], fn, *args, **kwargs):
    """Собрать все ошибки строки, а не падать на первой (как в отчёте легаси-импорта)."""
    try:
        return fn(*args, **kwargs)
    except _ImportRowError as exc:
        errors.append(str(exc))
        return None


def _import_number(raw: str, label: str) -> float | None:
    s = str(raw).replace(" ", "").replace(" ", "").replace(",", ".")
    if not s:
        return None
    try:
        return float(s)
    except ValueError as exc:
        raise _ImportRowError(f"{label}: «{raw}» — не число") from exc


def _import_int(raw: str, label: str) -> int | None:
    value = _import_number(raw, label)
    if value is None:
        return None
    if value < 0:
        raise _ImportRowError(f"{label}: значение не может быть отрицательным")
    return int(round(value))


def _import_kopecks(raw: str, label: str) -> int | None:
    value = _import_number(raw, label)
    if value is None:
        return None
    if value < 0:
        raise _ImportRowError(f"{label}: значение не может быть отрицательным")
    return int(round(value * 100))


def _import_dimension(raw: str, label: str) -> float:
    value = _import_number(raw, label)
    if value is None:
        return 0.0
    if value < 0:
        raise _ImportRowError(f"{label}: значение не может быть отрицательным")
    return float(value)


_IMPORT_TRUE = frozenset({"да", "true", "1", "истина", "yes", "y", "+"})
_IMPORT_FALSE = frozenset({"нет", "false", "0", "ложь", "no", "n", "-"})


def _import_bool(raw: str, label: str) -> bool:
    s = _fold_ci_str(raw).strip()
    if not s:
        return True
    if s in _IMPORT_TRUE:
        return True
    if s in _IMPORT_FALSE:
        return False
    raise _ImportRowError(f"{label}: укажите «да» или «нет»")


def _import_barcodes(raw: str) -> list[str]:
    out: list[str] = []
    for part in re.split(r"[;\n\r]+", str(raw)):
        code = _import_collapse_ws(part)
        if not code:
            continue
        if code in out:
            raise _ImportRowError(f"Штрих-код {code} указан в строке дважды")
        out.append(code)
    return out


def import_file_kind(filename: str) -> str | None:
    lower = str(filename).strip().lower()
    if lower.endswith(".xlsx") or lower.endswith(".xlsm"):
        return "xlsx"
    if lower.endswith(".xls"):
        return "xls"
    return None


def _import_header_map(header_cells: list[tuple[int, object]]) -> dict[int, str]:
    aliases: dict[str, str] = {}
    for key, title, extra in _IMPORT_COLUMNS:
        aliases[_fold_ci_str(title)] = key
        for alias in extra:
            aliases[_fold_ci_str(alias)] = key
    found: dict[int, str] = {}
    for col_idx, value in header_cells:
        key = aliases.get(_fold_ci_str(_import_cell_str(value)))
        if key and key not in found.values():
            found[col_idx] = key
    return found


def _import_rows_from_matrix(rows: list[list[object]]) -> list[dict]:
    """Общий разбор для xlsx и xls: поиск шапки, отбрасывание пустых строк, лимиты."""
    header_idx = -1
    col_map: dict[int, str] = {}
    # Шапка ищется в первых пяти строках: файл клиента часто приходит с титулом над таблицей.
    for idx, row in enumerate(rows[:5]):
        candidate = _import_header_map(list(enumerate(row)))
        if len(candidate) >= 3:
            header_idx = idx
            col_map = candidate
            break
    if not col_map:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"{_IMPORT_STRUCTURE_ERROR}: не найдена строка заголовков. Скачайте шаблон",
        )
    missing = [
        title for key, title, _ in _IMPORT_COLUMNS
        if key in _IMPORT_REQUIRED_COLUMNS and key not in col_map.values()
    ]
    if missing:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"{_IMPORT_STRUCTURE_ERROR}: нет обязательных колонок — " + ", ".join(missing),
        )

    out: list[dict] = []
    for idx in range(header_idx + 1, len(rows)):
        row = rows[idx]
        values = {key: "" for key, _, _ in _IMPORT_COLUMNS}
        has_value = False
        for col_idx, key in col_map.items():
            text = _import_cell_str(row[col_idx]) if col_idx < len(row) else ""
            values[key] = text
            if text:
                has_value = True
        if not has_value:
            continue
        values["row_no"] = idx + 1
        out.append(values)
        if len(out) > PRODUCT_IMPORT_MAX_ROWS:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail=f"В файле больше {PRODUCT_IMPORT_MAX_ROWS} строк — разбейте загрузку на части",
            )
    if not out:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Нет ни одной строки данных под заголовком",
        )
    return out


def _import_matrix_xlsx(content: bytes) -> list[list[object]]:
    from io import BytesIO

    from openpyxl import load_workbook

    try:
        wb = load_workbook(BytesIO(content), read_only=True, data_only=True)
    except Exception as exc:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail=_IMPORT_UNREADABLE_ERROR) from exc
    try:
        if not wb.sheetnames:
            raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail=_IMPORT_UNREADABLE_ERROR)
        ws = wb[wb.sheetnames[0]]
        return [list(row) for row in ws.iter_rows(values_only=True)]
    finally:
        wb.close()


def _import_matrix_xls(content: bytes) -> list[list[object]]:
    try:
        import xlrd
    except ImportError as exc:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail=_IMPORT_UNREADABLE_ERROR) from exc
    try:
        book = xlrd.open_workbook(file_contents=content)
        sheet = book.sheet_by_index(0)
    except Exception as exc:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail=_IMPORT_UNREADABLE_ERROR) from exc
    return [[sheet.cell_value(r, c) for c in range(sheet.ncols)] for r in range(sheet.nrows)]


def parse_product_import_workbook(content: bytes, *, filename: str) -> list[dict]:
    """Разбор файла в список сырых строк. В БД не обращается."""
    if not content or len(content) < 10:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail=_IMPORT_UNREADABLE_ERROR)
    if len(content) > PRODUCT_IMPORT_MAX_BYTES:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"Размер файла превышает допустимый лимит {PRODUCT_IMPORT_MAX_BYTES // (1024 * 1024)} МБ",
        )
    kind = import_file_kind(filename)
    if kind is None:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Неподдерживаемый формат файла. Нужен .xlsx, .xlsm или .xls",
        )
    # Расширение может врать — формат определяется сигнатурой (xlsx это zip).
    is_zip = content[0:2] == b"PK"
    rows = _import_matrix_xlsx(content) if is_zip else _import_matrix_xls(content)
    return _import_rows_from_matrix(rows)


def _import_dictionaries(connection: Any) -> tuple[dict[str, dict], dict[str, dict], dict[str, dict]]:
    types: dict[str, dict] = {}
    for r in connection.execute(
        "SELECT id, name, requires_color, requires_size FROM product_types "
        "WHERE is_active = 1 AND COALESCE(is_deleted, 0) = 0"
    ).fetchall():
        types[_fold_ci_str(_import_collapse_ws(r["name"]))] = {
            "id": str(r["id"]),
            "name": str(r["name"]),
            "requires_color": bool(r["requires_color"]),
            "requires_size": bool(r["requires_size"]),
        }
    colors = {
        _fold_ci_str(_import_collapse_ws(r["name"])): {"id": str(r["id"]), "name": str(r["name"])}
        for r in connection.execute(
            "SELECT id, name FROM colors WHERE is_active = 1 AND COALESCE(is_deleted, 0) = 0"
        ).fetchall()
    }
    sizes = {
        _fold_ci_str(_import_collapse_ws(r["name"])): {"id": str(r["id"]), "name": str(r["name"])}
        for r in connection.execute(
            "SELECT id, name FROM sizes WHERE is_active = 1 AND COALESCE(is_deleted, 0) = 0"
        ).fetchall()
    }
    return types, colors, sizes


def _parse_import_row(
    raw: dict, *, types: dict, colors: dict, sizes: dict
) -> tuple[dict | None, list[str], list[str]]:
    errors: list[str] = []
    warnings: list[str] = []

    name = _import_cell_str(raw.get("name"))
    if not name:
        errors.append("Укажите название товара")

    type_name = _import_cell_str(raw.get("type"))
    ptype: dict | None = None
    if not type_name:
        errors.append("Укажите тип товара")
    else:
        ptype = types.get(_fold_ci_str(type_name))
        if not ptype:
            errors.append(f"Тип товара «{type_name}» не найден в справочнике")

    color_name = _import_cell_str(raw.get("color"))
    color_id: str | None = None
    if color_name:
        color = colors.get(_fold_ci_str(color_name))
        if not color:
            errors.append(f"Цвет «{color_name}» не найден в справочнике")
        else:
            color_id = color["id"]
            if ptype and not ptype["requires_color"]:
                warnings.append("Тип товара не требует цвета — цвет всё равно будет записан")
    elif ptype and ptype["requires_color"]:
        errors.append("Для этого типа товара укажите цвет")

    size_name = _import_cell_str(raw.get("size"))
    size_id: str | None = None
    if ptype and ptype["requires_size"]:
        if not size_name:
            errors.append("Для этого типа товара укажите размер")
        else:
            size = sizes.get(_fold_ci_str(size_name))
            if not size:
                errors.append(f"Размер «{size_name}» не найден в справочнике")
            else:
                size_id = size["id"]
    elif size_name and ptype:
        warnings.append("Тип товара не требует размера — значение будет проигнорировано")

    length = _import_field(errors, _import_dimension, _import_cell_str(raw.get("length")), "Длина")
    width = _import_field(errors, _import_dimension, _import_cell_str(raw.get("width")), "Ширина")
    height = _import_field(errors, _import_dimension, _import_cell_str(raw.get("height")), "Высота")
    weight = _import_field(errors, _import_int, _import_cell_str(raw.get("weight")), "Вес")
    items_per_box = _import_field(errors, _import_int, _import_cell_str(raw.get("items_per_box")), "Кол-во в коробе")
    boxes_per_pallet = _import_field(errors, _import_int, _import_cell_str(raw.get("boxes_per_pallet")), "Коробов на палете")
    price_good = _import_field(errors, _import_kopecks, _import_cell_str(raw.get("price_good")), "Цена упаковки годный")
    price_defect = _import_field(errors, _import_kopecks, _import_cell_str(raw.get("price_defect")), "Цена упаковки брак")
    is_active = _import_field(errors, _import_bool, _import_cell_str(raw.get("active")), "Активен")
    barcodes = _import_field(errors, _import_barcodes, _import_cell_str(raw.get("barcodes")))

    sku = _import_cell_str(raw.get("sku"))
    if not sku:
        warnings.append("SKU не указан — товар будет заведён без SKU (уточнить позже)")
    if not (length or width or height):
        warnings.append("Габариты не заданы")
    if weight is None:
        warnings.append("Вес не указан")
    if items_per_box is None and boxes_per_pallet is None:
        warnings.append("Кратность коробов и палет не указана")

    if errors:
        return None, errors, warnings

    return {
        "sku": sku,
        "name": name,
        "type_id": ptype["id"],
        "type_name": ptype["name"],
        "requires_color": ptype["requires_color"],
        "requires_size": ptype["requires_size"],
        "color_id": color_id,
        "color_name": color_name if color_id else "",
        "size_id": size_id,
        "size_name": size_name if size_id else "",
        "length": length or 0.0,
        "width": width or 0.0,
        "height": height or 0.0,
        "weight_grams": weight,
        "items_per_box": items_per_box,
        "boxes_per_pallet": boxes_per_pallet,
        "price_good_kop": price_good,
        "price_defect_kop": price_defect,
        "is_active": True if is_active is None else is_active,
        "barcodes": barcodes or [],
    }, errors, warnings


_IMPORT_GROUP_ATTRS: tuple[tuple[str, str], ...] = (
    ("name", "название товара"),
    ("type_id", "тип товара"),
    ("weight_grams", "вес"),
    ("items_per_box", "кол-во в коробе"),
    ("boxes_per_pallet", "коробов на палете"),
    ("price_good_kop", "цену упаковки (годный)"),
    ("price_defect_kop", "цену упаковки (брак)"),
    ("is_active", "признак активности"),
)


def _merge_import_group_attrs(group: dict, parsed: dict) -> str | None:
    """Атрибуты товара берутся из первой строки группы; расхождение — ошибка строки."""
    for field, label in _IMPORT_GROUP_ATTRS:
        incoming = parsed[field]
        if incoming is None or incoming == "":
            continue
        current = group[field]
        if current is None or current == "":
            group[field] = incoming
            continue
        if current != incoming:
            return f"Расхождение с первой строкой товара (строка {group['first_row_no']}): {label}"
    return None


def _import_variant_identity(parsed: dict) -> tuple:
    if parsed["requires_size"]:
        return (parsed["color_id"] or "", parsed["size_id"] or "")
    return (
        parsed["color_id"] or "",
        round(float(parsed["length"]), 3),
        round(float(parsed["width"]), 3),
        round(float(parsed["height"]), 3),
    )


def _import_variant_sku(
    connection: Any,
    *,
    sku_base: str,
    client_id: str,
    color_label: str,
    size_label: str,
    requires_size: bool,
    multi_dim: bool,
    length: float,
    width: float,
    height: float,
    used: set[str],
) -> str:
    """SKU варианта по тем же правилам, что и при ручном заведении товара."""
    tokens: list[str] = []
    if color_label:
        tokens.append(_sku_token_from_label(color_label))
    if requires_size and size_label:
        tokens.append(_sku_token_from_label(size_label))
    if multi_dim and not requires_size:
        tokens.append(_sku_dim_token(length, width, height))
    candidate = "-".join([sku_base, *tokens]) if tokens else f"{sku_base}-1"
    sku = candidate
    salt = 0
    while sku in used or _variant_sku_in_use(connection, sku, None, client_id):
        salt += 1
        sku = f"{candidate}-{salt}"
    used.add(sku)
    return sku


def build_product_import_plan(connection: Any, *, client_id: str, rows: list[dict]) -> dict:
    """Проверить строки файла и собрать план записи. В БД ничего не пишет."""
    types, colors, sizes = _import_dictionaries(connection)
    groups: dict[tuple, dict] = {}
    out_rows: list[dict] = []
    barcode_rows: dict[str, int] = {}
    sku_case_rows: dict[str, tuple[str, int]] = {}

    for raw in rows:
        item = {
            "row_no": int(raw.get("row_no") or 0),
            "sku": _import_cell_str(raw.get("sku")),
            "name": _import_cell_str(raw.get("name")),
            "type_name": _import_cell_str(raw.get("type")),
            "color_name": _import_cell_str(raw.get("color")),
            "size_name": _import_cell_str(raw.get("size")),
            "variant_sku": "",
            "action": PRODUCT_IMPORT_ACTION_ERROR,
            "errors": [],
            "warnings": [],
        }
        out_rows.append(item)

        parsed, errors, warnings = _parse_import_row(raw, types=types, colors=colors, sizes=sizes)
        item["errors"].extend(errors)
        item["warnings"].extend(warnings)
        if parsed is None:
            continue

        for code in parsed["barcodes"]:
            if code in barcode_rows:
                item["errors"].append(f"Штрих-код {code} уже указан в строке {barcode_rows[code]}")
                continue
            owner = _find_barcode_owner(connection, code)
            if owner:
                item["errors"].append(f"Штрих-код {code} уже присвоен: {_barcode_owner_label(owner)}")
                continue
            barcode_rows[code] = item["row_no"]

        sku = parsed["sku"]
        if sku:
            folded = _fold_ci_str(sku)
            seen = sku_case_rows.get(folded)
            if seen and seen[0] != sku:
                item["errors"].append(
                    f"SKU {sku} отличается только регистром от строки {seen[1]} ({seen[0]})"
                )
            elif not seen:
                sku_case_rows[folded] = (sku, item["row_no"])

        if item["errors"]:
            continue

        key = ("sku", sku) if sku else ("pending", _fold_ci_str(parsed["name"]), parsed["type_id"])
        group = groups.get(key)
        if group is None:
            group = {
                "key": "|".join(str(k) for k in key),
                "sku": sku,
                "sku_pending": not sku,
                "first_row_no": item["row_no"],
                "requires_size": parsed["requires_size"],
                "existing_product_id": None,
                "variants": [],
                "identities": {},
            }
            for field, _ in _IMPORT_GROUP_ATTRS:
                group[field] = None
            groups[key] = group

        conflict = _merge_import_group_attrs(group, parsed)
        if conflict:
            item["errors"].append(conflict)
            continue

        identity = _import_variant_identity(parsed)
        if identity in group["identities"]:
            item["errors"].append(
                f"Такой вариант уже есть в файле (строка {group['identities'][identity]})"
            )
            continue
        group["identities"][identity] = item["row_no"]
        group["variants"].append({
            "row_no": item["row_no"],
            "color_id": parsed["color_id"],
            "color_name": parsed["color_name"],
            "size_id": parsed["size_id"],
            "size_name": parsed["size_name"],
            "length": parsed["length"],
            "width": parsed["width"],
            "height": parsed["height"],
            "barcodes": parsed["barcodes"],
            "identity": identity,
            "sku": "",
            "action": PRODUCT_IMPORT_ACTION_CREATE,
        })

    rows_by_no = {r["row_no"]: r for r in out_rows}
    plan_groups: list[dict] = []
    used_skus: set[str] = set()

    for group in groups.values():
        if not group["variants"]:
            continue
        first_row = rows_by_no.get(group["first_row_no"])
        group_error: str | None = None

        if group["sku"]:
            existing = connection.execute(
                "SELECT id, name, type_id FROM products WHERE client_id = ? AND sku = ? "
                "AND COALESCE(is_deleted, 0) = 0 AND COALESCE(sku_pending, 0) = 0",
                (client_id, group["sku"]),
            ).fetchone()
            if existing:
                if str(existing["type_id"]) != str(group["type_id"]):
                    group_error = "У товара с этим SKU в системе другой тип товара"
                else:
                    group["existing_product_id"] = str(existing["id"])
                    system_name = str(existing["name"] or "")
                    if _fold_ci_str(system_name) != _fold_ci_str(group["name"]) and first_row:
                        first_row["warnings"].append(
                            f"Название в системе отличается: {system_name} (останется прежним)"
                        )
            elif _sku_taken_for_client_except_product(connection, group["sku"], client_id, None):
                group_error = f"SKU {group['sku']} уже занят у этого клиента другим товаром"

        existing_identities: set[tuple] = set()
        if group["existing_product_id"]:
            for r in connection.execute(
                "SELECT color_id, size_id, length, width, height FROM product_variants "
                "WHERE product_id = ? AND COALESCE(is_deleted, 0) = 0",
                (group["existing_product_id"],),
            ).fetchall():
                if group["requires_size"]:
                    existing_identities.add((str(r["color_id"] or ""), str(r["size_id"] or "")))
                else:
                    existing_identities.add((
                        str(r["color_id"] or ""),
                        round(float(r["length"]), 3),
                        round(float(r["width"]), 3),
                        round(float(r["height"]), 3),
                    ))

        dims = {(v["length"], v["width"], v["height"]) for v in group["variants"]}
        multi_dim = len(dims) > 1
        for variant in group["variants"]:
            row = rows_by_no.get(variant["row_no"])
            if group_error:
                variant["action"] = PRODUCT_IMPORT_ACTION_ERROR
                if row:
                    row["errors"].append(group_error)
                continue
            if variant["identity"] in existing_identities:
                variant["action"] = PRODUCT_IMPORT_ACTION_SKIP
            elif group["existing_product_id"]:
                variant["action"] = PRODUCT_IMPORT_ACTION_APPEND
            else:
                variant["action"] = PRODUCT_IMPORT_ACTION_CREATE
            if variant["action"] != PRODUCT_IMPORT_ACTION_SKIP and not group["sku_pending"]:
                variant["sku"] = _import_variant_sku(
                    connection,
                    sku_base=group["sku"],
                    client_id=client_id,
                    color_label=variant["color_name"],
                    size_label=variant["size_name"],
                    requires_size=group["requires_size"],
                    multi_dim=multi_dim,
                    length=variant["length"],
                    width=variant["width"],
                    height=variant["height"],
                    used=used_skus,
                )
            if row:
                row["action"] = variant["action"]
                row["variant_sku"] = variant["sku"]
        if not group_error:
            group.pop("identities", None)
            for variant in group["variants"]:
                variant.pop("identity", None)
            plan_groups.append(group)

    rows_with_errors = sum(1 for r in out_rows if r["errors"])
    variants_new = sum(
        1 for g in plan_groups for v in g["variants"]
        if v["action"] in (PRODUCT_IMPORT_ACTION_CREATE, PRODUCT_IMPORT_ACTION_APPEND)
    )
    summary = {
        "rows_total": len(out_rows),
        "rows_ok": len(out_rows) - rows_with_errors,
        "rows_with_errors": rows_with_errors,
        "rows_with_warnings": sum(1 for r in out_rows if r["warnings"]),
        "products_new": sum(1 for g in plan_groups if not g["existing_product_id"]),
        "products_existing": sum(1 for g in plan_groups if g["existing_product_id"]),
        "variants_new": variants_new,
        "variants_skipped": sum(
            1 for g in plan_groups for v in g["variants"] if v["action"] == PRODUCT_IMPORT_ACTION_SKIP
        ),
        "barcodes_new": sum(
            len(v["barcodes"]) for g in plan_groups for v in g["variants"]
            if v["action"] != PRODUCT_IMPORT_ACTION_SKIP
        ),
        "import_ready": rows_with_errors == 0 and variants_new > 0,
        "can_import_partial": rows_with_errors > 0 and variants_new > 0,
    }
    return {"client_id": client_id, "groups": plan_groups, "rows": out_rows, "summary": summary}


def apply_product_import_plan(connection: Any, plan: dict, *, user_id: str) -> dict:
    """Записать план в БД (группы с ошибками в план не попадают). Коммитит вызывающий."""
    from config import INV_Q_DEFECT, INV_Q_GOOD
    from modules.pricing.service import add_price
    from modules.timesheet.service import business_today

    client_id = str(plan["client_id"])
    now = _now()
    effective_from = business_today().isoformat()

    for group in plan["groups"]:
        product_id = group.get("existing_product_id")
        is_active = 0 if group["is_active"] is False else 1
        if not product_id:
            product_id = str(uuid4())
            connection.execute(
                """
                INSERT INTO products (id, name, type_id, client_id, supplier_id, sku, sku_pending,
                                      weight_grams, items_per_box, boxes_per_pallet,
                                      image_url, gallery_json, is_active, created_at, creator_id)
                VALUES (?, ?, ?, ?, NULL, ?, ?, ?, ?, ?, NULL, NULL, ?, ?, ?)
                """,
                (
                    product_id, group["name"], group["type_id"], client_id,
                    group["sku"] or "", 1 if group["sku_pending"] else 0,
                    group["weight_grams"], group["items_per_box"], group["boxes_per_pallet"],
                    is_active, now, user_id,
                ),
            )
            # Тариф пишется только для нового товара: у существующего цена ведётся
            # в справочнике «Стоимость упаковки», и файл не должен её переписывать.
            if group["price_good_kop"] is not None:
                add_price(connection, product_id=product_id, client_id=client_id, quality=INV_Q_GOOD,
                          price_kop=group["price_good_kop"], effective_from=effective_from, user_id=user_id)
            if group["price_defect_kop"] is not None:
                add_price(connection, product_id=product_id, client_id=client_id, quality=INV_Q_DEFECT,
                          price_kop=group["price_defect_kop"], effective_from=effective_from, user_id=user_id)

        for variant in group["variants"]:
            if variant["action"] == PRODUCT_IMPORT_ACTION_SKIP:
                continue
            variant_id = str(uuid4())
            connection.execute(
                """
                INSERT INTO product_variants (id, product_id, client_id, color_id, size_id,
                                              length, width, height, sku, sku_pending,
                                              images_json, is_active, created_at, is_deleted)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, '[]', ?, ?, 0)
                """,
                (
                    variant_id, product_id, client_id, variant["color_id"], variant["size_id"],
                    float(variant["length"]), float(variant["width"]), float(variant["height"]),
                    variant["sku"], 1 if group["sku_pending"] else 0, is_active, now,
                ),
            )
            for code in variant["barcodes"]:
                connection.execute(
                    "INSERT INTO product_barcodes (id, product_id, variant_id, barcode, source, created_at, created_by, is_deleted) "
                    "VALUES (?,?,?,?,?,?,?,0)",
                    (str(uuid4()), product_id, variant_id, code, _IMPORT_BARCODE_SOURCE, now, user_id),
                )
    return plan["summary"]


def build_product_import_template_bytes(connection: Any) -> bytes:
    """xlsx-шаблон: шапка, строка-пример, лист «Справочники» со списками этого инстанса."""
    from io import BytesIO

    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.datavalidation import DataValidation

    wb = Workbook()
    ws = wb.active
    ws.title = "Товары"
    titles = [title for _, title, _ in _IMPORT_COLUMNS]
    ws.append(titles)
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(fill_type="solid", start_color="305496")
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 22
    ws.freeze_panes = "A2"

    widths = [18, 34, 20, 18, 12, 11, 11, 11, 11, 16, 20, 26, 20, 18, 11]
    for idx, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width

    types_rows = connection.execute(
        "SELECT name FROM product_types WHERE is_active = 1 AND COALESCE(is_deleted, 0) = 0 ORDER BY LOWER(name)"
    ).fetchall()
    colors_rows = connection.execute(
        "SELECT name FROM colors WHERE is_active = 1 AND COALESCE(is_deleted, 0) = 0 ORDER BY LOWER(name)"
    ).fetchall()
    sizes_rows = connection.execute(
        "SELECT name FROM sizes WHERE is_active = 1 AND COALESCE(is_deleted, 0) = 0 ORDER BY LOWER(name)"
    ).fetchall()

    ws.append([
        "BASE-001",
        "ДАННЫЕ ДЛЯ ПРИМЕРА, УДАЛИТЕ СТРОКУ",
        str(types_rows[0]["name"]) if types_rows else "Футболка",
        str(colors_rows[0]["name"]) if colors_rows else "Черный",
        str(sizes_rows[0]["name"]) if sizes_rows else "44",
        30, 20, 5, 250, 20, 12, "4600000000001;4600000000002", "12,50", "5,00", "да",
    ])

    ref = wb.create_sheet("Справочники")
    ref.append(["Типы товара", "Цвета", "Размеры"])
    for cell in ref[1]:
        cell.font = Font(bold=True)
    ref.column_dimensions["A"].width = 28
    ref.column_dimensions["B"].width = 28
    ref.column_dimensions["C"].width = 18
    for idx in range(max(len(types_rows), len(colors_rows), len(sizes_rows))):
        ref.cell(row=idx + 2, column=1, value=str(types_rows[idx]["name"]) if idx < len(types_rows) else None)
        ref.cell(row=idx + 2, column=2, value=str(colors_rows[idx]["name"]) if idx < len(colors_rows) else None)
        ref.cell(row=idx + 2, column=3, value=str(sizes_rows[idx]["name"]) if idx < len(sizes_rows) else None)

    for col_idx, ref_col, rows_len in (("C", "A", len(types_rows)), ("D", "B", len(colors_rows)), ("E", "C", len(sizes_rows))):
        if not rows_len:
            continue
        dv = DataValidation(
            type="list",
            formula1=f"='Справочники'!${ref_col}$2:${ref_col}${rows_len + 1}",
            allow_blank=True,
        )
        ws.add_data_validation(dv)
        dv.add(f"{col_idx}2:{col_idx}1000")

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def build_product_import_report_bytes(plan: dict) -> bytes:
    """Отчёт проверки: исходные колонки + «Ошибки» и «Предупреждения» — правь и грузи заново."""
    from io import BytesIO

    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "Проверка"
    headers = ["Строка", "SKU базовый", "Название товара", "Тип товара", "Цвет", "Размер",
               "SKU варианта", "Действие", "Ошибки", "Предупреждения"]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)
    ws.freeze_panes = "A2"
    for idx, width in enumerate([9, 18, 34, 20, 18, 12, 22, 16, 60, 60], start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width

    error_fill = PatternFill(fill_type="solid", start_color="FCE4E4")
    for row in plan["rows"]:
        ws.append([
            row["row_no"], row["sku"], row["name"], row["type_name"], row["color_name"],
            row["size_name"], row["variant_sku"], PRODUCT_IMPORT_ACTION_LABELS.get(row["action"], row["action"]),
            "; ".join(row["errors"]), "; ".join(row["warnings"]),
        ])
        if row["errors"]:
            for cell in ws[ws.max_row]:
                cell.fill = error_fill

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def purge_stale_product_import_batches(
    connection: Any, *, older_than_hours: int = PRODUCT_IMPORT_BATCH_TTL_HOURS
) -> None:
    """Удалить непринятые превью загрузок. Применённые пакеты остаются как аудит."""
    from datetime import UTC, datetime, timedelta

    cutoff = (datetime.now(UTC) - timedelta(hours=older_than_hours)).isoformat()
    connection.execute(
        "DELETE FROM product_import_batches WHERE status = ? AND created_at < ?",
        (PRODUCT_IMPORT_STATUS_PREVIEW, cutoff),
    )


class _Code(NamedTuple):
    barcode: str
    source: str | None
    store_id: str | None


def _label_candidates(codes: list[_Code], store_id: str | None) -> list[_Code]:
    """Из чего выбирают этикетку строки: сначала коды её магазина, потом общие.

    Магазин — доменный признак, а не украшение: у варианта в кабинете Ozon и в
    кабинете WB разные ШК, и клеить чужой нельзя. Когда магазин строки не указан,
    кандидаты — общие коды, а если и таких нет, остаются все.
    """
    own = [c for c in codes if store_id and c.store_id == store_id]
    if own:
        return own
    general = [c for c in codes if not c.store_id]
    return general or codes


def build_barcode_labels(
    connection: Any, items: list[BarcodeLabelRequestItem], *, all_codes: bool = False
) -> BarcodeLabelsResponse:
    """Печатные этикетки ШК по вариантам: код рисуется здесь и нигде не хранится.

    Картинка — чистая функция от цифр, поэтому сохранённый файл только протухал бы
    после смены цвето-размера варианта или прихода нового кода из кабинета МП.
    Загруженная вручную этикетка (`product_barcode_files`) этим не отменяется:
    она остаётся приоритетной там, где менеджер её приложил.
    """
    product_ids = list({str(it.product_id) for it in items})
    placeholders = ",".join("?" for _ in product_ids)
    variant_rows = connection.execute(
        f"""
        SELECT v.id, v.product_id, v.sku, v.color_id, v.size_id,
               col.name AS color_name, sz.name AS size_name,
               p.name AS product_name
        FROM product_variants v
        JOIN products p ON p.id = v.product_id
        LEFT JOIN colors col ON col.id = v.color_id
        LEFT JOIN sizes sz ON sz.id = v.size_id
        WHERE v.product_id IN ({placeholders})
          AND COALESCE(v.is_deleted, 0) = 0
          AND COALESCE(p.is_deleted, 0) = 0
        """,
        product_ids,
    ).fetchall()
    by_identity: dict[tuple[str, str, str], Any] = {}
    for r in variant_rows:
        key = (str(r["product_id"]), str(r["color_id"] or ""), str(r["size_id"] or ""))
        by_identity.setdefault(key, r)

    barcode_rows = connection.execute(
        f"""
        SELECT product_id, variant_id, barcode, source, store_id
        FROM product_barcodes
        WHERE product_id IN ({placeholders}) AND COALESCE(is_deleted, 0) = 0
        ORDER BY created_at
        """,
        product_ids,
    ).fetchall()
    by_variant: dict[str, list[_Code]] = {}
    product_level: dict[str, list[_Code]] = {}
    for r in barcode_rows:
        code = _Code(
            barcode=str(r["barcode"]),
            source=str(r["source"]) if r["source"] else None,
            store_id=str(r["store_id"]) if r["store_id"] else None,
        )
        if r["variant_id"]:
            by_variant.setdefault(str(r["variant_id"]), []).append(code)
        else:
            product_level.setdefault(str(r["product_id"]), []).append(code)

    labels: list[BarcodeLabelItem] = []
    missing: list[BarcodeLabelMissingItem] = []
    for it in items:
        pid = str(it.product_id)
        row = by_identity.get((pid, str(it.color_id or ""), str(it.size_id or "")))
        if row is None:
            missing.append(BarcodeLabelMissingItem(
                product_id=pid, color_id=str(it.color_id) if it.color_id else None,
                size_id=str(it.size_id) if it.size_id else None,
                label=pid, reason="Вариант не найден",
            ))
            continue
        vid = str(row["id"])
        attrs = " / ".join(x for x in [row["color_name"], row["size_name"]] if x)
        label = str(row["sku"]) + (f" ({attrs})" if attrs else "")
        # Код без варианта — легаси карточек с единственным вариантом: он опознаёт
        # тот же товар, и печатать по нему можно.
        codes = by_variant.get(vid) or product_level.get(pid) or []
        if not codes:
            missing.append(BarcodeLabelMissingItem(
                product_id=pid, color_id=str(it.color_id) if it.color_id else None,
                size_id=str(it.size_id) if it.size_id else None,
                variant_id=vid, label=label, reason="У варианта нет штрих-кода",
            ))
            continue
        candidates = _label_candidates(codes, str(it.store_id or "") or None)
        mixed = len({(c.store_id or "", c.source or "") for c in candidates}) > 1
        pinned = str(it.barcode or "").strip()
        if pinned and pinned not in [c.barcode for c in codes]:
            # Выбранный на строке код сняли с карточки. Молча подставить другой нельзя:
            # это ровно та ошибка, ради которой выбор и заводился.
            missing.append(BarcodeLabelMissingItem(
                product_id=pid, color_id=str(it.color_id) if it.color_id else None,
                size_id=str(it.size_id) if it.size_id else None,
                variant_id=vid, label=label,
                reason="Выбранный код снят с карточки — выберите заново",
            ))
            continue
        if pinned:
            chosen = [c for c in codes if c.barcode == pinned]
        elif all_codes:
            # Шторка выбора показывает все коды варианта, кандидаты — первыми.
            chosen = candidates + [c for c in codes if c not in candidates]
        else:
            chosen = candidates[:1]
        for code in chosen:
            svg, modules = _barcode_svg(code.barcode)
            labels.append(BarcodeLabelItem(
                product_id=pid,
                color_id=str(it.color_id) if it.color_id else None,
                size_id=str(it.size_id) if it.size_id else None,
                variant_id=vid,
                barcode=code.barcode,
                barcode_svg=svg,
                modules=modules,
                product_name=str(row["product_name"] or ""),
                sku=str(row["sku"]),
                color_name=row["color_name"],
                size_name=row["size_name"],
                qty=int(it.qty),
                source=code.source,
                store_id=code.store_id,
                chosen=bool(pinned) and code.barcode == pinned,
                barcode_count=len(candidates),
                mixed_origin=mixed,
            ))
    return BarcodeLabelsResponse(items=labels, missing=missing)
